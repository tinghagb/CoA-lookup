#!/usr/bin/env python3
"""
BioLegend Certificate of Analysis Lookup Tool
==============================================
• Scan a 2D barcode → look up CoA → populate a spreadsheet
• OR upload an existing inventory spreadsheet → batch-fill all CoA fields

Usage:
    pip install -r requirements.txt
    python app.py
    Open http://localhost:5050
"""

import io
import re
import json
import time
import queue
import logging
import threading
from typing import Optional, Generator

import requests
import cloudscraper
from bs4 import BeautifulSoup
from flask import Flask, request, jsonify, send_file, Response
import openpyxl
from openpyxl.styles import PatternFill, Font

logging.basicConfig(level=logging.INFO, format="%(levelname)s  %(message)s")
logger = logging.getLogger(__name__)

app = Flask(__name__)

# ════════════════════════════════════════════════════════════════════
# CONSTANTS
# ════════════════════════════════════════════════════════════════════

# BioLegend fluorochrome list — ordered longest/most-specific first
# so that "PerCP/Cyanine5.5" matches before "PerCP"
FLUORS = [
    "Brilliant Violet 421", "Brilliant Violet 480", "Brilliant Violet 510",
    "Brilliant Violet 570", "Brilliant Violet 605", "Brilliant Violet 650",
    "Brilliant Violet 711", "Brilliant Violet 750", "Brilliant Violet 785",
    "Brilliant Violet 786",
    "BV421", "BV480", "BV510", "BV570", "BV605", "BV650", "BV711", "BV785", "BV786",
    "Brilliant Blue 515", "Brilliant Blue 700", "BB515", "BB700",
    "Brilliant Ultraviolet 395", "Brilliant Ultraviolet 496", "Brilliant Ultraviolet 563",
    "Brilliant Ultraviolet 615", "Brilliant Ultraviolet 661", "Brilliant Ultraviolet 737",
    "Brilliant Ultraviolet 805",
    "BUV395", "BUV496", "BUV563", "BUV615", "BUV661", "BUV737", "BUV805",
    "PerCP/Cyanine5.5", "PerCP-Cyanine5.5", "PerCP/Cy5.5", "PerCP-Cy5.5", "PerCP",
    "PE/Cyanine5", "PE/Cyanine7", "PE-Cy5", "PE-Cy7",
    "PE/Dazzle594", "PE-Dazzle594",
    "PE/Texas Red", "PE-Texas Red",
    "PE-CF594",
    "PE",
    "APC/Cyanine7", "APC-Cy7", "APC/Fire750", "APC-Fire750",
    "APC-R700", "APC-R750",
    "APC",
    "Alexa Fluor 488", "Alexa Fluor 532", "Alexa Fluor 594",
    "Alexa Fluor 647", "Alexa Fluor 700", "Alexa Fluor 750",
    "AF488", "AF647", "AF700",
    "FITC",
    "Pacific Blue", "Pacific Orange",
    "Zombie Aqua", "Zombie Red", "Zombie Yellow", "Zombie Green", "Zombie NIR",
    "LIVE/DEAD",
    "7-AAD",
    "DAPI",
    "eFluor 450", "eFluor 506", "eFluor 660", "eFluor 780",
    "Super Bright 436", "Super Bright 600", "Super Bright 645", "Super Bright 702",
    "SB436", "SB600", "SB645", "SB702",
    "Fire 810",
]

SPECIES_PREFIXES = re.compile(
    r"^Anti[-\s]+"
    r"(?:Human(?:/Mouse)?(?:/Rat)?|Mouse(?:/Human)?(?:/Rat)?|Rat(?:/Mouse)?|"
    r"Hamster|Rabbit|Monkey|Canine|Feline|Equine|Porcine|Bovine|"
    r"Human/Mouse/Rat|Mouse/Human/Rat|Mouse/Rat)(?:/Human)?\s+",
    re.IGNORECASE,
)

SPECIES_SUFFIXES = re.compile(
    r"\s+Antibody$|\s+Kit$|\s+Reagent$|\s+Cocktail.*$", re.IGNORECASE
)

# ════════════════════════════════════════════════════════════════════
# BARCODE PARSER
# ════════════════════════════════════════════════════════════════════

def parse_barcode(raw: str, vendor: str = "") -> dict:
    """
    Parse a scanned barcode string, optionally biased by the selected vendor.
    Supported vendors: "biolegend", "bd", "miltenyi" (case-insensitive).
    If vendor is blank we auto-detect.
    """
    raw = raw.strip()
    vendor = (vendor or "").strip().lower()
    result = {"raw": raw, "catalog_no": "", "lot_no": "", "scan_date": "",
              "format": "unknown", "vendor": vendor}
    if not raw:
        return result

    # ── Vendor-specific parsers ─────────────────────────────────────────────
    # Miltenyi: long digit string with GS1 AI 91 (company-internal data).
    # Across all observed samples the structure is:
    #   positions 0-1   = "91"   (AI 91 — Miltenyi-internal)
    #   positions 2-4   = 3-digit sub-code (varies: e.g. 656, 906, 746, 956)
    #   positions 5-13  = 9-digit catalog number  (e.g. 130123710 → "130-123-710")
    #   positions 14-23 = 10-digit lot number     (e.g. 5241004962)
    #   positions 24+   = trailer (additional GS1 AIs / padding)
    # Matches any barcode beginning with "91" and at least 24 digits long.
    def _parse_miltenyi(s: str) -> bool:
        digits = re.sub(r"\D", "", s)
        # Must start with AI 91 and have enough digits for cat + lot
        if not digits.startswith("91") or len(digits) < 24:
            return False
        cat = digits[5:14]       # 9-digit catalog
        lot = digits[14:24]      # 10-digit lot
        # Sanity: cat must look like a Miltenyi catalog (starts with 130 most
        # of the time; relax to "all digits" for forward compat).
        if not cat.isdigit() or not lot.isdigit():
            return False
        result.update(catalog_no=cat, lot_no=lot,
                      format="miltenyi_gs1", vendor="miltenyi")
        return True

    # BD: GS1 DataMatrix, no parentheses, AI 01/17/10/240
    #   "01" + 14-digit GTIN + "17" + YYMMDD + "10" + 7-digit lot + ...
    def _parse_bd(s: str) -> bool:
        digits = re.sub(r"\D", "", s)
        if not digits.startswith("01") or len(digits) < 16 + 8 + 9:
            return False
        gtin = digits[2:16]
        pos = 16
        # AI 17 (expiry) — optional
        if digits[pos:pos + 2] == "17" and len(digits) >= pos + 8:
            d = digits[pos + 2: pos + 8]
            result["scan_date"] = f"20{d[0:2]}/{d[2:4]}/{d[4:6]}"
            pos += 8
        # AI 10 (lot, fixed 7 for BD)
        if digits[pos:pos + 2] != "10":
            return False
        lot = digits[pos + 2: pos + 9]
        # BD catalog = GTIN digits 8-13 (1-based) = slice [7:13]
        cat = str(int(gtin[7:13]))
        result.update(catalog_no=cat, lot_no=lot,
                      format="bd_gs1", vendor="bd")
        return True

    # BioLegend standard: "344742 B402098 2026/03/05 04:16:50"
    def _parse_biolegend_standard(s: str) -> bool:
        parts = s.split()
        if len(parts) >= 2 and re.match(r"^\d{4,8}$", parts[0]) and re.match(r"^[A-Za-z]\d{4,}$", parts[1], re.I):
            result.update(catalog_no=parts[0], lot_no=parts[1].upper(),
                          format="biolegend_standard", vendor="biolegend")
            if len(parts) > 2:
                result["scan_date"] = " ".join(parts[2:])
            return True
        return False

    # GS1 DataMatrix WITH parenthesized AIs (rare — human-readable)
    def _parse_gs1_parens(s: str) -> bool:
        if not re.search(r"\(\d{2}\)", s):
            return False
        result["format"] = "gs1_parenthesized"
        m = re.search(r"\(10\)([^(]+)", s)
        if m: result["lot_no"] = m.group(1).strip().upper()
        m = re.search(r"\(01\)(\d{13,14})", s)
        if m: result["catalog_no"] = str(int(m.group(1)[7:13]))
        m = re.search(r"\(17\)(\d{6})", s)
        if m:
            d = m.group(1)
            result["scan_date"] = f"20{d[0:2]}/{d[2:4]}/{d[4:6]}"
        return True

    # ── Dispatch — vendor-biased first, then autodetect ───────────────────
    if vendor == "miltenyi" and _parse_miltenyi(raw):
        return result
    if vendor == "bd" and _parse_bd(raw):
        return result
    if vendor == "biolegend":
        if _parse_biolegend_standard(raw): return result
        if _parse_gs1_parens(raw):         return result

    # Autodetect (used when vendor is blank or vendor-specific failed)
    if _parse_miltenyi(raw):            return result
    if _parse_biolegend_standard(raw):  return result
    if _parse_gs1_parens(raw):          return result
    if _parse_bd(raw):                  return result

    # Tab-separated
    if "\t" in raw:
        p = [x.strip() for x in raw.split("\t")]
        result.update(catalog_no=p[0], lot_no=p[1].upper() if len(p) > 1 else "",
                      format="tab_separated")
        return result

    # Lot only / Catalog only
    if re.match(r"^[A-Za-z]\d{4,}$", raw):
        result.update(lot_no=raw.upper(), format="lot_only")
    elif re.match(r"^\d{4,8}$", raw):
        result.update(catalog_no=raw, format="catalog_only")
    else:
        result.update(lot_no=raw, format="unknown_as_lot")
    return result


# ════════════════════════════════════════════════════════════════════
# MARKER / FLUOR EXTRACTOR
# ════════════════════════════════════════════════════════════════════

# Finnish month names → English (BioLegend CoA dates are locale-dependent)
_FINNISH_MONTHS = {
    "tammikuuta": "January",  "helmikuuta": "February", "maaliskuuta": "March",
    "huhtikuuta": "April",    "toukokuuta": "May",       "kesäkuuta":   "June",
    "heinäkuuta": "July",     "elokuuta":   "August",    "syyskuuta":   "September",
    "lokakuuta":  "October",  "marraskuuta":"November",  "joulukuuta":  "December",
}

def _normalize_date(s: str) -> str:
    """Translate Finnish (or other locale) month names to English,
    and strip any trailing ISO-style time fragment (e.g. ' T', ' 00:00:00')."""
    for fi, en in _FINNISH_MONTHS.items():
        s = s.replace(fi, en)
    # Remove trailing ' T' or time portion like 'T12:00:00' or ' 00:00:00'
    s = re.sub(r"\s+T$", "", s)
    s = re.sub(r"\s+T\d{2}:\d{2}.*$", "", s)
    s = re.sub(r"\s+\d{2}:\d{2}:\d{2}.*$", "", s)
    return s.strip()


_ANTI_SPECIES = re.compile(
    r"(?:Anti[-\s]+)?"
    r"(?:Human(?:/Mouse)?(?:/Rat)?|Mouse(?:/Human)?(?:/Rat)?|Rat(?:/Mouse)?|"
    r"Hamster|Rabbit|Monkey|Canine|Feline|Porcine|Bovine|"
    r"Human/Mouse/Rat|Mouse/Human/Rat)\s*",
    re.IGNORECASE,
)


def extract_marker_fluor(product_name: str) -> tuple[str, str]:
    """
    Parse a BioLegend product name in either format:
      • "Anti-Human CD3 PerCP/Cyanine5.5 Antibody"   (new-style)
      • "Brilliant Violet 605™ anti-human CD8"        (CoA-page style)
    Returns (marker, fluor).
    """
    if not product_name:
        return "", ""

    name = product_name.strip()
    # Strip trademark symbols for matching, but keep original for display
    name_clean = re.sub(r"[™®©]", "", name).strip()

    # Detect fluorochrome
    fluor = ""
    for f in FLUORS:
        if re.search(re.escape(f), name_clean, re.IGNORECASE):
            fluor = f
            break

    # Remove the fluor text entirely
    working = name_clean
    if fluor:
        working = re.sub(re.escape(fluor), "", working, flags=re.IGNORECASE).strip()

    # Remove "Anti-[Species]" and bare species names
    working = SPECIES_PREFIXES.sub("", working)
    working = _ANTI_SPECIES.sub("", working)

    # Remove trailing "Antibody", "Kit", "Reagent", etc.
    working = SPECIES_SUFFIXES.sub("", working).strip()

    # Strip orphaned "anti-" or "anti " prefix not consumed by species removal
    # e.g. "anti-BrdU" after removing fluor/species → "BrdU"
    working = re.sub(r"^[Aa]nti[-\s]+", "", working).strip()

    # Strip leftover punctuation/symbols
    working = re.sub(r"^[\s,\-/]+|[\s,\-/]+$", "", working).strip()

    return working, fluor


# ════════════════════════════════════════════════════════════════════
# HTTP SESSION
# ════════════════════════════════════════════════════════════════════

# The BioLegend CoA form lives at this fixed Kentico CMS page.
# Confirmed via DevTools: <form action="/Default.aspx?ID=44925&action=Detail" method="post">
# Submit button:  name="Coa.Search.Submit"  value="Coa.Search.Submit"
COA_FORM_URL = "https://www.biolegend.com/Default.aspx?ID=44925&action=Detail"

# cloudscraper mimics a real browser's TLS fingerprint and JS-challenge handling,
# which prevents 403 bot-blocking from BioLegend / Cloudflare.
_SESSION = cloudscraper.create_scraper(
    browser={"browser": "chrome", "platform": "windows", "mobile": False}
)
_SESSION.headers.update({
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://www.biolegend.com/",
})


def _get(url: str, params: dict = None, timeout: int = 20) -> Optional[requests.Response]:
    try:
        r = _SESSION.get(url, params=params, timeout=timeout, allow_redirects=True)
        logger.info(f"GET {r.url} → {r.status_code} ({len(r.text):,} chars)")
        return r if r.ok else None
    except Exception as e:
        logger.warning(f"Request error ({url}): {e}")
        return None


def _post(url: str, data: dict, timeout: int = 20) -> Optional[requests.Response]:
    try:
        r = _SESSION.post(url, data=data, timeout=timeout, allow_redirects=True)
        logger.info(f"POST {r.url} → {r.status_code} ({len(r.text):,} chars)")
        return r if r.ok else None
    except Exception as e:
        logger.warning(f"POST error ({url}): {e}")
        return None


# ════════════════════════════════════════════════════════════════════
# COA FETCHER
# ════════════════════════════════════════════════════════════════════

# Prefer lxml (fast) but fall back to Python's stdlib html.parser if it
# isn't installed — keeps Windows installs working even when lxml fails
# to compile or its pre-built wheel is unavailable.
try:
    import lxml  # noqa: F401
    _BS_PARSER = "lxml"
except Exception:
    _BS_PARSER = "html.parser"
    logger.warning("lxml not available — using html.parser fallback")


def _get_aspnet_hidden_fields(html: str) -> dict:
    """Extract ASP.NET hidden fields (__VIEWSTATE etc.) from a page."""
    soup = BeautifulSoup(html, _BS_PARSER)
    fields = {}
    for inp in soup.find_all("input", type="hidden"):
        name = inp.get("name", "")
        if name:
            fields[name] = inp.get("value", "")
    return fields


def _find_lot_input_name(html: str) -> str:
    """
    Find the name attribute of the lot number text input on the CoA form.
    Falls back to the Kentico-style name pattern.
    """
    soup = BeautifulSoup(html, _BS_PARSER)
    # Look for a text input whose name or id contains 'lot' (case-insensitive)
    for inp in soup.find_all("input", type=["text", "search", None]):
        name = inp.get("name", "") or ""
        id_  = inp.get("id",   "") or ""
        if re.search(r"lot", name, re.IGNORECASE) or re.search(r"lot", id_, re.IGNORECASE):
            return name
    # Kentico widget naming convention: matches "Coa.Search.Submit" button pattern
    # → input likely named "Coa.Search.LotNo" or "txtLotNo"
    for candidate in ("Coa.Search.LotNo", "txtLotNo", "LotNo", "lotNo"):
        if candidate in html:
            return candidate
    return "Coa.Search.LotNo"   # best guess


def _empty_coa(vendor: str, catalog_no: str, lot_no: str, fallback_url: str = "") -> dict:
    return {
        "found": False, "source_url": "",
        "vendor": vendor,
        "product_name": "", "catalog_no": catalog_no, "lot_no": lot_no,
        "marker": "", "fluor": "",
        "clone": "", "host_species": "", "reactivity": "", "isotype": "",
        "formulation": "", "storage": "", "expiry_date": "", "concentration": "",
        "volume": "", "applications": "",
        "optimal_dilution": "",
        "tests": [], "specs": {},
        "error": "",
        "fallback_url": fallback_url,
    }


def fetch_coa(vendor: str, catalog_no: str, lot_no: str) -> dict:
    """Vendor-aware dispatcher."""
    v = (vendor or "biolegend").strip().lower()
    if v == "bd":
        return fetch_bd_coa(catalog_no, lot_no)
    if v == "miltenyi":
        return fetch_miltenyi_coa(catalog_no, lot_no)
    return fetch_biolegend_coa(catalog_no, lot_no)


def fetch_biolegend_coa(catalog_no: str, lot_no: str) -> dict:
    """
    Fetch CoA from BioLegend by POSTing to the Kentico form at Default.aspx?ID=44925.
    Falls back to old PHP URL and plain GET attempts if POST fails.
    """
    base = _empty_coa("biolegend", catalog_no, lot_no, fallback_url=COA_FORM_URL)

    if not lot_no:
        base["error"] = "Lot number is required to look up a CoA."
        return base

    # ── Strategy 1: POST to the confirmed Kentico form URL ────────────────────
    r_form = _get(COA_FORM_URL)
    if r_form:
        hidden = _get_aspnet_hidden_fields(r_form.text)
        lot_field = _find_lot_input_name(r_form.text)
        post_data = {
            **hidden,
            lot_field: lot_no,
            "Coa.Search.Submit": "Coa.Search.Submit",
        }
        r = _post(COA_FORM_URL, post_data)
        if r:
            parsed = _parse_coa_page(r.text, catalog_no, lot_no)
            if parsed["found"]:
                parsed["source_url"] = r.url
                base.update(parsed)
                logger.info(f"CoA found via POST form for {catalog_no}/{lot_no}")
                return base

    # ── Strategy 2: Old PHP redirect (may land on Default.aspx result) ────────
    r = _get("https://www.biolegend.com/certificate_of_analysis.php", {"lot": lot_no})
    if r:
        parsed = _parse_coa_page(r.text, catalog_no, lot_no)
        if parsed["found"]:
            parsed["source_url"] = r.url
            base.update(parsed)
            logger.info(f"CoA found via old-PHP for {catalog_no}/{lot_no}")
            return base

    base["error"] = (
        "Could not retrieve CoA automatically. "
        "Click the link below and enter the lot number manually."
    )
    return base


def _parse_coa_page(html: str, catalog_no: str, lot_no: str) -> dict:
    return _parse_html(html, catalog_no, lot_no)


def _parse_html(html: str, catalog_no: str, lot_no: str) -> dict:
    """
    Extract structured CoA fields from BioLegend's Default.aspx CoA page.

    The confirmed page structure (from DevTools) is:
      <h2>  Product name  </h2>
      Catalog No. XXXXXX
      Lot No. XXXXXX
      <b>Clone:</b>            SK1
      <b>Format:</b>           Brilliant Violet 605™   ← this is the Fluor
      <b>Isotype:</b>          Mouse IgG1, κ
      <b>Concentration:</b>    50 µg/mL
      <b>Volume:</b>           500 µL
      <b>Applications:</b>     FC - Quality tested
      <b>Storage:</b>          ...
      <b>Formulation:</b>      ...
      <b>Expiration Date:</b>  syyskuuta 29, 2025
    """
    soup = BeautifulSoup(html, _BS_PARSER)
    for tag in soup(["script", "style", "nav", "footer", "noscript"]):
        tag.decompose()

    body = soup.get_text(separator=" ", strip=True)

    result = {
        "found": False, "product_name": "",
        "catalog_no": catalog_no, "lot_no": lot_no,
        "marker": "", "fluor": "",
        "clone": "", "host_species": "", "reactivity": "", "isotype": "",
        "formulation": "", "storage": "", "expiry_date": "", "concentration": "",
        "volume": "", "applications": "",
        "optimal_dilution": "",
        "tests": [], "specs": {},
    }

    # Bail on error/empty pages
    bad = ["no certificate", "not found", "invalid lot", "no results",
           "does not exist", "no coa available"]
    if any(p in body.lower() for p in bad) and len(body) < 5_000:
        return result

    # Must contain "Certificate of Analysis" to be a real CoA page
    if "certificate of analysis" not in body.lower():
        return result

    # ── Product name ─────────────────────────────────────────────────────────
    # DOM traversal is unreliable because BioLegend's sidebar/nav HTML appears
    # *before* the main CoA content in source order (e.g. "Follow Us",
    # "Worldwide Ordering" are picked up by find_all_previous).
    #
    # Text-layer strategy: extract the plain-text of the page with newlines,
    # then find the line(s) sandwiched between "Certificate of Analysis" and
    # "Catalog No. <catalog_no>". The product name is always the LAST non-junk
    # line in that sandwich.
    _BAD_NAME = re.compile(
        r"certificate|catalog|lot\s+no|biolegend|revvity|cookie|privacy|"
        r"javascript|loading|please wait|follow us|worldwide|sign in|"
        r"my account|contact us|about us|resources|careers|"
        r"popular links|quick links|ordering|shop\b",
        re.IGNORECASE,
    )

    _page_lines = soup.get_text(separator="\n", strip=True)

    def _find_product_name_in_text(text: str, cat: str) -> str:
        # Try: text between "Certificate of Analysis" and "Catalog No. <cat>"
        patterns = []
        if cat:
            patterns.append(
                rf"Certificate of Analysis(.*?)Catalog\s+No\.?\s*{re.escape(cat)}"
            )
        # Broader: just between the two anchors
        patterns.append(
            r"Certificate of Analysis(.*?)Catalog\s+No\.?"
        )
        for pat in patterns:
            m = re.search(pat, text, re.IGNORECASE | re.DOTALL)
            if not m:
                continue
            between = m.group(1)
            lines = [l.strip() for l in between.split("\n") if l.strip()]
            # Walk lines from BOTTOM (closest to "Catalog No.") upward
            for line in reversed(lines):
                if not (6 < len(line) < 250):
                    continue
                if line.endswith(":"):
                    continue
                if _BAD_NAME.search(line):
                    continue
                return line
        return ""

    pname = _find_product_name_in_text(_page_lines, catalog_no)
    if pname:
        result["product_name"] = pname
        result["found"] = True

    # ── Catalog / Lot from "Catalog No. XXXX" / "Lot No. XXXX" text ──────────
    m = re.search(r"Catalog\s+No\.?\s*(\d{4,8})", body)
    if m and not result["catalog_no"]:
        result["catalog_no"] = m.group(1)

    m = re.search(r"Lot\s+No\.?\s*([A-Za-z]\d{4,})", body)
    if m and not result["lot_no"]:
        result["lot_no"] = m.group(1).upper()
        result["found"] = True

    # ── Field label → result key mapping ─────────────────────────────────────
    # "Format" on BioLegend CoA = the fluorochrome/conjugate → Fluor column
    LABEL_FIELD_MAP = {
        "clone":            "clone",
        "format":           "fluor",        # ← BioLegend "Format" = fluorochrome
        "isotype":          "isotype",
        "concentration":    "concentration",
        "volume":           "volume",
        "applications":     "applications",
        "storage":          "storage",
        "formulation":      "formulation",
        "preparation":      "specs",
        "expiration date":  "expiry_date",
        "expiry date":      "expiry_date",
        "host species":     "host_species",
        "reactivity":       "reactivity",
    }

    def _assign_field(label_raw: str, value_raw: str):
        label = label_raw.strip().rstrip(":").lower()
        value = value_raw.strip()
        if not label or not value:
            return
        field = LABEL_FIELD_MAP.get(label)
        if field == "specs":
            result["specs"][label_raw.strip().rstrip(":")] = value
        elif field and not result.get(field):
            result[field] = value
            result["found"] = True

    # ── Strategy A: parse table.certificateAnalysis directly ─────────────────
    # Confirmed from DevTools: the CoA data is in a <table class="certificateAnalysis">
    # Structure: <tr><td><b>Label:</b></td><td>Value</td></tr>
    coa_table = (
        soup.find("table", class_="certificateAnalysis")
        or soup.find("table", class_=re.compile(r"certificate", re.I))
    )
    if coa_table:
        for row in coa_table.find_all("tr"):
            cells = row.find_all(["td", "th"])
            if len(cells) >= 2:
                _assign_field(cells[0].get_text(strip=True), cells[1].get_text(strip=True))

    # ── Strategy B: any <b> tag whose value is in the adjacent <td> ──────────
    # Handles: <td><b>Format:</b></td><td>Brilliant Violet 605™</td>
    for b_tag in soup.find_all("b"):
        label = b_tag.get_text(strip=True).rstrip(":").lower()
        if label not in LABEL_FIELD_MAP:
            continue
        value = ""
        # Check adjacent <td> first (table-row structure)
        parent_td = b_tag.find_parent("td")
        if parent_td:
            next_td = parent_td.find_next_sibling("td")
            if next_td:
                value = next_td.get_text(strip=True)
        # Fallback: text/element siblings of <b> itself (inline structure)
        if not value:
            parts = []
            for sib in b_tag.next_siblings:
                if getattr(sib, "name", None) == "b":
                    break
                text = sib.strip() if isinstance(sib, str) else sib.get_text(strip=True)
                if text:
                    parts.append(text)
            value = " ".join(parts).strip()
        if value:
            _assign_field(b_tag.get_text(strip=True), value)

    # ── Normalize date (Finnish months → English) ─────────────────────────────
    if result["expiry_date"]:
        result["expiry_date"] = _normalize_date(result["expiry_date"])

    # ── Derive Marker + Fluor from product name ───────────────────────────────
    # Also try the "Format" field value directly as the fluor (most reliable)
    if result.get("fluor"):
        result["fluor"] = re.sub(r"[™®©]", "", result["fluor"]).strip()

    if result["product_name"]:
        marker, fluor_from_name = extract_marker_fluor(result["product_name"])
        if not result["marker"] and marker:
            result["marker"] = marker
        if not result["fluor"] and fluor_from_name:
            result["fluor"] = fluor_from_name

    # ── Fallback: regex on body text ──────────────────────────────────────────
    regex_fb = {
        "clone":         r"\bClone[:\s]+([A-Z0-9/\-\.]+)",
        "isotype":       r"\bIsotype[:\s]+([A-Za-z0-9 ,κλ]+?)(?:\s{2,}|\n|<)",
        "concentration": r"\bConcentration[:\s]+([0-9.]+\s*(?:mg/mL|µg/mL|μg/mL|ug/mL))",
        "expiry_date":   r"\bExpiration Date[:\s]+([^\n<]{4,20})",
        "storage":       r"\bStorage[:\s]+([^\n<]{10,80})",
    }
    for field, pat in regex_fb.items():
        if not result.get(field):
            m = re.search(pat, body)
            if m:
                val = m.group(1).strip()
                if field == "expiry_date":
                    val = _normalize_date(val)
                result[field] = val

    return result


# ════════════════════════════════════════════════════════════════════
# BD COA FETCHER
# ════════════════════════════════════════════════════════════════════
#
# BD's QualityCert page uses a 3-step internal JSON API (discovered via
# Chrome DevTools → Network):
#
#   1. POST /regulatory/auth/token     → {"guid":"<client-guid>"}  → JWT
#   2. POST /regulatory/api/search/qc  → {qualityCerts:{qualityCert:[
#                                          {materialNumber,batchNumber}]}…}
#                                       Authorization: Bearer <JWT>
#                                       → {data:[{id:"ecc:cat~lot~hash",…}]}
#   3. POST /regulatory/api/viewPDF    → {"fileID":"ecc:cat~lot~hash"}
#                                       Authorization: Bearer <JWT>
#                                       → application/pdf bytes
#
# The client-guid below was captured live; it matches the JWT's `kid`
# header, indicating it identifies a stable signing key (not a per-session
# token). If BD ever rotates it, you'd need to scrape it from the SPA's
# JS bundle.

BD_QC_FORM_URL     = "https://regdocs.bd.com/regdocs/qcinfo"
BD_QC_RESULTS_URL  = "https://regdocs.bd.com/regdocs/qcSearchResults"   # legacy, fallback only
BD_TOKEN_URL       = "https://regdocs.bd.com/regulatory/auth/token"
BD_SEARCH_QC_URL   = "https://regdocs.bd.com/regulatory/api/search/qc"
BD_VIEW_PDF_URL    = "https://regdocs.bd.com/regulatory/api/viewPDF"
BD_CLIENT_GUID     = "4eb0d0e0-b23b-4654-b005-3c80a84d3f8e"

# Cached JWT — short-lived (≈15 min), so we lazily refresh on 401.
_BD_TOKEN_CACHE = {"jwt": "", "expires": 0.0}


def _download_pdf(url: str, params: dict = None,
                  data: dict = None, method: str = "GET",
                  timeout: int = 30) -> Optional[bytes]:
    """Try to download a PDF document. Returns bytes if response is a PDF."""
    try:
        if method == "POST":
            r = _SESSION.post(url, data=data, params=params,
                              timeout=timeout, allow_redirects=True)
        else:
            r = _SESSION.get(url, params=params,
                             timeout=timeout, allow_redirects=True)
        ctype = r.headers.get("Content-Type", "").lower()
        logger.info(f"{method} {r.url} → {r.status_code} "
                    f"({len(r.content):,} bytes, {ctype})")
        if not r.ok:
            return None
        # Accept either application/pdf or any body that begins with %PDF
        if "pdf" in ctype or r.content[:5] == b"%PDF-":
            return r.content
        return None
    except Exception as e:
        logger.warning(f"PDF download error ({url}): {e}")
        return None


def _pdf_text(pdf_bytes: bytes) -> str:
    """Extract all text from a PDF using pypdf."""
    try:
        from pypdf import PdfReader
    except ImportError:
        logger.error("pypdf not installed")
        return ""
    try:
        reader = PdfReader(io.BytesIO(pdf_bytes))
        return "\n".join((page.extract_text() or "") for page in reader.pages)
    except Exception as e:
        logger.warning(f"pypdf extraction error: {e}")
        return ""


def _bd_common_headers(token: str = "") -> dict:
    """Standard browser-like headers required by regdocs.bd.com APIs."""
    h = {
        "accept": "application/json, text/plain, */*",
        "accept-language": "en-US,en;q=0.9",
        "content-type": "application/json",
        "origin": "https://regdocs.bd.com",
        "referer": "https://regdocs.bd.com/regdocs/qcinfo",
        "sec-fetch-dest": "empty",
        "sec-fetch-mode": "cors",
        "sec-fetch-site": "same-origin",
    }
    if token:
        h["authorization"] = f"Bearer {token}"
    return h


def _bd_extract_jwt_from_any(payload) -> str:
    """Walk a JSON-decoded payload looking for any JWT-shaped string."""
    if isinstance(payload, str):
        s = payload.strip().strip('"')
        if s.startswith("ey") and s.count(".") == 2 and len(s) > 50:
            return s
        return ""
    if isinstance(payload, dict):
        # Try common field names first
        for k in ("token", "access_token", "jwt", "authToken", "auth_token",
                  "Token", "AccessToken", "JWT", "AuthToken", "data", "result"):
            if k in payload:
                got = _bd_extract_jwt_from_any(payload[k])
                if got: return got
        for v in payload.values():
            got = _bd_extract_jwt_from_any(v)
            if got: return got
    elif isinstance(payload, list):
        for item in payload:
            got = _bd_extract_jwt_from_any(item)
            if got: return got
    return ""


_BD_WARMED_UP = {"v": False}

def _bd_warmup_session():
    """One-time GET to BD's QC page so that any session cookies are set."""
    if _BD_WARMED_UP["v"]:
        return
    try:
        r = _SESSION.get(BD_QC_FORM_URL, timeout=20, allow_redirects=True)
        logger.info(f"BD warmup GET {BD_QC_FORM_URL} → {r.status_code}")
        _BD_WARMED_UP["v"] = True
    except Exception as e:
        logger.warning(f"BD warmup error: {e}")


def _bd_get_token(force_refresh: bool = False) -> tuple:
    """
    Fetch a fresh JWT for the BD QC API. Cached for ~13 min.
    Returns (token, error_msg). On success error_msg is "".
    """
    now = time.time()
    if (not force_refresh
            and _BD_TOKEN_CACHE["jwt"]
            and _BD_TOKEN_CACHE["expires"] > now + 30):
        return _BD_TOKEN_CACHE["jwt"], ""

    # Warm up the session so cookies are set
    _bd_warmup_session()

    try:
        r = _SESSION.post(
            BD_TOKEN_URL,
            headers=_bd_common_headers(),
            json={"guid": BD_CLIENT_GUID},
            timeout=30,
        )
        snippet = (r.text or "")[:300].replace("\n", " ")
        logger.info(f"BD token POST → {r.status_code} "
                    f"({len(r.content)} bytes) "
                    f"ct={r.headers.get('content-type','')}")
        logger.info(f"BD token body[:300]: {snippet}")

        if r.status_code != 200:
            return "", f"BD token endpoint returned HTTP {r.status_code}: {snippet}"

        token = ""
        # Try JSON first (irrespective of content-type — BD sometimes mislabels)
        try:
            data = r.json()
            token = _bd_extract_jwt_from_any(data)
        except Exception:
            data = None

        # Fall back to raw text if it looks like a JWT
        if not token:
            txt = (r.text or "").strip().strip('"')
            if txt.startswith("ey") and txt.count(".") == 2:
                token = txt

        if not token:
            return "", (f"BD token response shape unexpected. "
                        f"Status={r.status_code}, body[:200]={snippet[:200]}")

        # JWT exp claim: assume ~15 min; cache for 13 min to be safe.
        _BD_TOKEN_CACHE["jwt"] = token
        _BD_TOKEN_CACHE["expires"] = now + 13 * 60
        return token, ""
    except Exception as e:
        return "", f"BD token network error: {e!r}"


def _bd_search_qc(token: str, catalog_no: str, lot_no: str) -> dict:
    """
    Call BD's QC search endpoint. Returns the full JSON response or {} on error.
    Expected response contains a `data` array whose items have:
        id              = "ecc:<materialNumber>~<batchNumber>~<32-hex-hash>"
        materialNumber  = catalog
        batch           = lot
        documentId      = the 32-hex-hash portion
    """
    body = {
        "qualityCerts": {
            "qualityCert": [{"materialNumber": catalog_no, "batchNumber": lot_no}]
        },
        "webDisplayLang": "en",
        "internalSearch": False,
    }
    try:
        r = _SESSION.post(
            BD_SEARCH_QC_URL,
            headers=_bd_common_headers(token),
            json=body,
            timeout=30,
        )
        logger.info(f"BD search/qc POST → {r.status_code}")
        if r.status_code == 401:
            return {"_unauthorized": True}
        r.raise_for_status()
        return r.json()
    except Exception as e:
        logger.warning(f"BD search/qc error: {e}")
        return {}


def _bd_walk_for_file_id(node) -> str:
    """Recursively search a JSON tree for the first 'ecc:'-prefixed id."""
    if isinstance(node, dict):
        for key in ("id", "fileID", "fileId"):
            v = node.get(key)
            if isinstance(v, str) and v.startswith("ecc:"):
                return v
        for v in node.values():
            r = _bd_walk_for_file_id(v)
            if r:
                return r
    elif isinstance(node, list):
        for item in node:
            r = _bd_walk_for_file_id(item)
            if r:
                return r
    return ""


def _bd_view_pdf(token: str, file_id: str) -> Optional[bytes]:
    """POST fileID → return raw PDF bytes (or None on failure)."""
    try:
        r = _SESSION.post(
            BD_VIEW_PDF_URL,
            headers=_bd_common_headers(token),
            json={"fileID": file_id},
            timeout=60,
        )
        ctype = r.headers.get("content-type", "").lower()
        logger.info(f"BD viewPDF POST → {r.status_code} "
                    f"({len(r.content):,} bytes, {ctype})")
        if r.status_code == 401:
            return b"__UNAUTHORIZED__"
        if not r.ok:
            return None
        # Path A: response IS the PDF
        if "pdf" in ctype or r.content[:5] == b"%PDF-":
            return r.content
        # Path B: JSON wrapping a base64 blob or redirect URL
        try:
            data = r.json()
            if isinstance(data, dict):
                if isinstance(data.get("url"), str):
                    r2 = _SESSION.get(data["url"], timeout=60)
                    if r2.ok and (r2.content[:5] == b"%PDF-"
                                  or "pdf" in r2.headers.get("content-type", "").lower()):
                        return r2.content
                blob = data.get("pdf") or data.get("data") or data.get("file")
                if isinstance(blob, str) and len(blob) > 100:
                    import base64
                    try:
                        decoded = base64.b64decode(blob, validate=False)
                        if decoded[:5] == b"%PDF-":
                            return decoded
                    except Exception:
                        pass
        except Exception:
            pass
        return None
    except Exception as e:
        logger.warning(f"BD viewPDF error: {e}")
        return None


def fetch_bd_coa(catalog_no: str, lot_no: str) -> dict:
    """
    Fetch a BD QualityCert PDF using the regdocs.bd.com 3-step API:
      token → search/qc → viewPDF → parse.
    """
    base = _empty_coa("bd", catalog_no, lot_no, fallback_url=BD_QC_FORM_URL)

    if not (catalog_no and lot_no):
        base["error"] = "Both catalog and lot are required for BD lookups."
        return base

    # ── Step 1: token ────────────────────────────────────────────────────
    token, terr = _bd_get_token()
    if not token:
        base["error"] = (f"Could not obtain BD auth token. {terr} "
                         "Click the link below to look up manually.")
        return base

    # ── Step 2: search/qc to get fileID ──────────────────────────────────
    search = _bd_search_qc(token, catalog_no, lot_no)
    if isinstance(search, dict) and search.get("_unauthorized"):
        # token expired — refresh and retry once
        token, terr = _bd_get_token(force_refresh=True)
        if not token:
            base["error"] = f"BD auth token refresh failed. {terr}"
            return base
        search = _bd_search_qc(token, catalog_no, lot_no)

    file_id = _bd_walk_for_file_id(search)
    if not file_id:
        base["error"] = (f"BD search returned no QC document for "
                         f"catalog {catalog_no}, lot {lot_no}.")
        return base

    # ── Step 3: viewPDF → bytes ──────────────────────────────────────────
    pdf_bytes = _bd_view_pdf(token, file_id)
    if pdf_bytes == b"__UNAUTHORIZED__":
        token, _ = _bd_get_token(force_refresh=True)
        pdf_bytes = _bd_view_pdf(token, file_id) if token else None
    if not pdf_bytes or pdf_bytes == b"__UNAUTHORIZED__":
        base["error"] = ("BD viewPDF call failed. "
                         "Click the link below to look up manually.")
        return base

    base["source_url"] = f"{BD_VIEW_PDF_URL}#{file_id}"

    # ── Step 4: parse ────────────────────────────────────────────────────
    parsed = _parse_bd_pdf(pdf_bytes, catalog_no, lot_no)
    base.update(parsed)
    return base


def _parse_bd_pdf(pdf_bytes: bytes, catalog_no: str, lot_no: str) -> dict:
    """
    Parse a BD QualityCert PDF.
      • Product Name:    e.g. "Hu CD16 R718 3G8 100Tst"
        → Marker  = CD16
        → Fluor   = R718
        → Clone   = 3G8
        → Species = Hu  → "Human"
      • Catalog Number:  566969
      • Batch Number:    5050874  (= Lot)
      • Expiration Date: 2027/07/31
      • Manufacture Date
      • Optimal Conc:    1.000 uG/Test
    """
    out = {"found": False}
    text = _pdf_text(pdf_bytes)
    if not text:
        return out

    # ── Field extraction ──────────────────────────────────────────────────
    grab = lambda pat: (m.group(1).strip() if (m := re.search(pat, text, re.IGNORECASE)) else "")

    product = grab(r"Product\s*Name[:\s]+([^\n\r]+)")
    cat     = grab(r"Catalog\s*(?:Number|No\.?|#)[:\s]+(\S+)")
    lot     = grab(r"Batch\s*(?:Number|No\.?|#)[:\s]+(\S+)") \
              or grab(r"Lot\s*(?:Number|No\.?|#)[:\s]+(\S+)")
    # Expiration Date often followed by underscores/line-art – stop before any "_"
    expiry  = grab(r"Expir(?:ation|y)\s*Date\s*[:\-]?\s*([0-9/\-]+)")
    mfg     = grab(r"Manufacture(?:d)?\s*Date\s*[:\-]?\s*[-\s]*([0-9/\-]+)")
    # BD concentration format: "<Unit>   -<Value>" (hyphens are column dividers).
    # Example: "Human FC Optimal Conc uG/Test           -1.000__________"
    conc_raw = grab(r"Optimal\s*Conc(?:entration|\.)?\s*([^\n\r]+)") \
               or grab(r"Concentration\s*([^\n\r]+)")
    conc = _clean_bd_concentration(conc_raw)
    isotype = grab(r"Isotype[:\s]+([^\n\r]+)")
    appl    = grab(r"Applications?[:\s]+([^\n\r]+)")
    storage = grab(r"Storage[:\s]+([^\n\r]+)")

    if product or cat or lot:
        out["found"] = True

    if product:
        out["product_name"] = product
        marker, fluor, clone, species = _parse_bd_product_name(product)
        if marker:  out["marker"]       = marker
        if fluor:   out["fluor"]        = fluor
        if clone:   out["clone"]        = clone
        if species: out["host_species"] = species

    # Preserve input cat/lot — only fill from PDF if input was empty
    if cat and not catalog_no: out["catalog_no"] = cat
    if lot and not lot_no:     out["lot_no"]     = lot
    if expiry:  out["expiry_date"]   = expiry.strip("_-/").replace("-", "/")
    if mfg:     out["specs"]         = {"Manufacture Date": mfg.strip("_-")}
    if conc:    out["concentration"] = conc
    if isotype: out["isotype"]       = isotype
    if appl:    out["applications"]  = appl
    if storage: out["storage"]       = storage

    return out


def _clean_bd_concentration(raw: str) -> str:
    """
    Convert BD's table-layout concentration string into a clean value.
    Examples:
      "uG/Test           -1.000__________"          -> "1.000 uG/Test"
      "1.000 uG/Test"                               -> "1.000 uG/Test"
      "Human FC Optimal Conc uG/Test           -1.000"
                                                    -> "1.000 uG/Test"
    """
    if not raw:
        return ""
    s = re.sub(r"_{2,}", "", raw).strip()
    # Find <unit>  <dash><value>  pattern anywhere in the string
    m = re.search(r"([µuμ][gG]/\s*\S+|[mM][gG]/\s*\S+|ng/\s*\S+|%\S*)"
                  r"\s*[-\s]\s*([0-9]+\.?[0-9]*)", s)
    if m:
        return f"{m.group(2).strip()} {m.group(1).strip()}"
    # Value-first pattern
    m = re.search(r"([0-9]+\.?[0-9]*)\s*([µuμmn]?[gG]/\S+|%\S*)", s)
    if m:
        return f"{m.group(1)} {m.group(2)}"
    return s.strip(" -_")


# Map BD species prefixes → readable names
_BD_SPECIES = {
    "Hu":  "Human",  "Mu":  "Mouse",   "Rt":  "Rat",
    "Hum": "Human",  "Mou": "Mouse",   "Hu/Mu": "Human/Mouse",
    "NHP": "Non-human Primate",
}

# BD fluor / format tokens — tried in order, longest first
_BD_FLUORS = [
    "BB515", "BB700", "BB755", "BB790",
    "BUV395", "BUV496", "BUV563", "BUV615", "BUV661", "BUV737", "BUV805",
    "BV421", "BV480", "BV510", "BV570", "BV605", "BV650", "BV711", "BV750", "BV786",
    "PerCP-Cy5.5", "PerCP-Cy5", "PerCP",
    "PE-Cy7", "PE-Cy5.5", "PE-Cy5", "PE-CF594",
    "APC-H7", "APC-Cy7", "APC-R700", "APC",
    "Alexa Fluor 647", "Alexa Fluor 700", "Alexa Fluor 488", "AF488", "AF647", "AF700",
    "FITC", "PE",
    "R718", "R670", "R660",
    "Pacific Blue", "PacBlue",
    "V450", "V500",
]


def _parse_bd_product_name(name: str) -> tuple[str, str, str, str]:
    """
    BD product names: 'Hu CD16 R718 3G8 100Tst'
                       species marker fluor clone size
    Returns (marker, fluor, clone, species).
    """
    marker = fluor = clone = species = ""
    if not name:
        return marker, fluor, clone, species

    # 1. Species prefix
    parts = name.split()
    if parts and parts[0] in _BD_SPECIES:
        species = _BD_SPECIES[parts[0]]
        parts = parts[1:]

    # 2. Marker — first token starting with letter or 'CD'
    if parts:
        marker = parts[0]
        parts = parts[1:]

    # 3. Fluor — match against known BD fluor tokens
    name_clean = " ".join(parts)
    for f in _BD_FLUORS:
        if re.search(r"\b" + re.escape(f) + r"\b", name_clean, re.IGNORECASE):
            fluor = f
            name_clean = re.sub(r"\b" + re.escape(f) + r"\b",
                                "", name_clean, flags=re.IGNORECASE).strip()
            break

    # 4. Clone — typically a short alphanumeric token (not a size)
    for tok in name_clean.split():
        if re.match(r"^\d+(Tst|Test|Tests?|µg|ug|mg)$", tok, re.IGNORECASE):
            continue   # size — skip
        if re.match(r"^[A-Za-z0-9.\-/]{2,12}$", tok):
            clone = tok
            break

    return marker, fluor, clone, species


# ════════════════════════════════════════════════════════════════════
# MILTENYI COA FETCHER
# ════════════════════════════════════════════════════════════════════

MILTENYI_COA_BASE = "https://assets.miltenyibiotec.com/coa"


def _miltenyi_catalog_variants(catalog_no: str) -> list:
    """
    Generate every plausible Miltenyi catalog formatting we should try in the
    CoA URL. The canonical Miltenyi catalog is a 9-digit number rendered as
    "XXX-XXX-XXX" (e.g. 130-123-710). We always try the dashed form first
    (it's what assets.miltenyibiotec.com publishes), then fall back to raw
    digits and other plausible shapes.
    """
    cat = (catalog_no or "").strip()
    digits = re.sub(r"\D", "", cat)
    out = []

    def _add(v):
        if v and v not in out:
            out.append(v)

    if len(digits) == 9:
        # Canonical 9-digit Miltenyi product code → XXX-XXX-XXX (preferred)
        _add(f"{digits[0:3]}-{digits[3:6]}-{digits[6:9]}")
        _add(digits)                                   # raw 9-digit fallback
    elif len(digits) == 10:
        # 10-digit: probably 9-digit catalog + 1 check/version digit. Try the
        # 9-digit dashed form first (drop the trailing digit), then variants.
        _add(f"{digits[0:3]}-{digits[3:6]}-{digits[6:9]}")      # 3-3-3 (drop last)
        _add(digits[0:9])                                       # raw 9-digit
        _add(f"{digits[0:3]}-{digits[3:6]}-{digits[6:10]}")     # 3-3-4
        _add(f"{digits[0:3]}-{digits[3:7]}-{digits[7:10]}")     # 3-4-3
        _add(f"{digits[0:3]}-{digits[3:6]}-{digits[6:9]}-{digits[9]}")  # 3-3-3-1
        _add(digits)                                            # raw 10-digit
    elif len(digits) == 11:
        _add(f"{digits[0:3]}-{digits[3:6]}-{digits[6:11]}")
        _add(digits)

    # User-typed form (preserves dashes when manually entered as "130-123-710")
    _add(cat)
    # Bare digits as a last resort
    _add(digits)

    return out


def _miltenyi_url(catalog_no: str, lot_no: str, cat_variant: str = "") -> str:
    """Format a Miltenyi CoA URL using a particular catalog variant."""
    use = cat_variant.strip() if cat_variant else catalog_no.strip()
    return f"{MILTENYI_COA_BASE}/CoA_{use}_{lot_no.strip()}.pdf"


def fetch_miltenyi_coa(catalog_no: str, lot_no: str) -> dict:
    """
    Miltenyi CoAs are published at deterministic URLs:
      https://assets.miltenyibiotec.com/coa/CoA_<catalog>_<lot>.pdf
    Catalog formatting varies (with/without dashes, 9 vs. 10 digits) so we
    try every plausible variant before giving up.
    """
    base = _empty_coa("miltenyi", catalog_no, lot_no,
                      fallback_url="https://www.miltenyibiotec.com/US-en/resources/"
                                   "technical-documents/certificates.html")

    if not (catalog_no and lot_no):
        base["error"] = "Both catalog and lot are required for Miltenyi lookups."
        return base

    pdf_bytes = None
    used_url = ""
    tried = []
    for variant in _miltenyi_catalog_variants(catalog_no):
        url = _miltenyi_url(catalog_no, lot_no, cat_variant=variant)
        tried.append(url)
        pdf_bytes = _download_pdf(url)
        if pdf_bytes:
            used_url = url
            break

    if not pdf_bytes:
        base["error"] = ("Could not retrieve Miltenyi CoA. Tried "
                         f"{len(tried)} URL variant(s). "
                         "Click the link below to look up manually.")
        base["fallback_url"] = tried[0] if tried else ""
        # Log the URLs we tried for diagnostics
        for u in tried:
            logger.info(f"Miltenyi tried: {u}")
        return base

    base["source_url"] = used_url
    parsed = _parse_miltenyi_pdf(pdf_bytes, catalog_no, lot_no)
    base.update(parsed)
    return base


def _parse_miltenyi_pdf(pdf_bytes: bytes, catalog_no: str, lot_no: str) -> dict:
    """Parse a Miltenyi CoA PDF. Field positions vary across products."""
    out = {"found": False}
    text = _pdf_text(pdf_bytes)
    if not text:
        return out

    grab = lambda pat: (m.group(1).strip()
                        if (m := re.search(pat, text, re.IGNORECASE)) else "")

    # Product name — Miltenyi labels it "Name:" on the CoA PDF.
    # Try "Name:" first, then fall back to other vendor variants.
    product = grab(r"(?:^|\n)\s*Name\s*[:\-]\s*([^\n\r]+)") \
           or grab(r"Product\s*Name[:\s]+([^\n\r]+)") \
           or grab(r"Product[:\s]+([^\n\r]+)") \
           or grab(r"Description[:\s]+([^\n\r]+)") \
           or grab(r"Antibody[:\s]+([^\n\r]+)")

    cat = grab(r"Order\s*(?:no|number|code)[:\s\.]+(\S+)") \
       or grab(r"Cat(?:alog(?:ue)?)?\.?\s*(?:no|number|#)[:\s\.]+(\S+)")
    lot = grab(r"Lot[:\s\.]+(\S+)") \
       or grab(r"Batch[:\s\.]+(\S+)")

    # Expiry — Miltenyi writes "Expiration date: 29 Dec 2026".
    # Grab the whole rest of the line, then pull a clean date out of it
    # (this is robust to multiple spaces, non-breaking whitespace, etc.).
    expiry_line = grab(r"Expir(?:ation|y)\s*(?:date)?[:\s\-]+([^\n\r]+)") \
               or grab(r"Use\s*by[:\s\-]+([^\n\r]+)")
    expiry = _extract_date_from_line(expiry_line)

    clone   = grab(r"Clone[:\s]+(\S+)")
    isotype = grab(r"Isotype[:\s]+([^\n\r]+)")
    conc    = grab(r"Concentration[:\s]+([^\n\r]+)")
    storage = grab(r"Storage[:\s]+([^\n\r]+)")

    if product or cat or lot:
        out["found"] = True

    if product:
        out["product_name"] = product
        marker, fluor = extract_marker_fluor(product)
        if marker: out["marker"] = marker
        if fluor:  out["fluor"]  = fluor

    # CRITICAL: do NOT overwrite catalog/lot that came from the barcode or
    # manual entry. The PDF text-extraction can produce mangled tokens, and
    # the input values are already known-good. Only fill in if input is blank.
    if cat and not catalog_no: out["catalog_no"] = cat
    if lot and not lot_no:     out["lot_no"]     = lot

    if expiry:  out["expiry_date"]   = _normalize_miltenyi_date(expiry)
    if clone:   out["clone"]         = clone
    if isotype: out["isotype"]       = isotype
    if conc:    out["concentration"] = conc
    if storage: out["storage"]       = storage

    return out


def _extract_date_from_line(line: str) -> str:
    """
    Pull a date out of a free-form line.
    Tries (in order):
      • '29 Dec 2026'   — day  month-name  year
      • 'Dec 29, 2026'  — month-name  day, year
      • '2026-12-29'    — ISO numeric
      • '29/12/2026'    — DMY numeric
    Returns '' if nothing matches.
    """
    if not line:
        return ""
    s = line.strip()
    patterns = [
        r"\d{1,2}\s+[A-Za-z]{3,9}\.?\s+\d{2,4}",        # 29 Dec 2026
        r"[A-Za-z]{3,9}\.?\s+\d{1,2},?\s+\d{2,4}",      # Dec 29, 2026
        r"\d{4}[\-/.]\d{1,2}[\-/.]\d{1,2}",              # 2026-12-29
        r"\d{1,2}[\-/.]\d{1,2}[\-/.]\d{2,4}",            # 29/12/2026
    ]
    for pat in patterns:
        m = re.search(pat, s)
        if m:
            return re.sub(r"\s+", " ", m.group(0)).strip()
    return ""


def _normalize_miltenyi_date(raw: str) -> str:
    """
    Preserve textual Miltenyi dates (e.g. '29 Dec 2026') verbatim.
    Only normalize purely numeric formats by switching dashes/dots to slashes.
    """
    s = (raw or "").strip()
    if not s:
        return ""
    # If any letter is present, keep it as-is (textual month form)
    if re.search(r"[A-Za-z]", s):
        return re.sub(r"\s+", " ", s)
    return s.replace("-", "/").replace(".", "/")


# ════════════════════════════════════════════════════════════════════
# SPREADSHEET PROCESSOR
# ════════════════════════════════════════════════════════════════════

# Column header → CoA field name mapping
COL_FIELD_MAP = {
    "marker":           "marker",
    "fluor":            "fluor",
    "fluorochrome":     "fluor",
    "clone":            "clone",
    "isotype":          "isotype",
    "concentration":    "concentration",
    "expiration date":  "expiry_date",
    "expiry date":      "expiry_date",
    "expiry":           "expiry_date",
    "optimal dilution": "optimal_dilution",
    "dilution":         "optimal_dilution",
}

def _find_col(headers: list, *keywords: str) -> Optional[int]:
    """Return 1-based column index for the first header matching any keyword."""
    for kw in keywords:
        for i, h in enumerate(headers):
            if kw.lower() in str(h).lower():
                return i + 1  # 1-based
    return None


def process_spreadsheet(file_bytes: bytes, progress_q: queue.Queue,
                        default_vendor: str = "biolegend") -> bytes:
    """
    Load xlsx, look up CoA for each row that has catalog+lot but missing fields,
    fill in the data, and return the updated xlsx bytes.
    `default_vendor` is used when the spreadsheet has no Vendor column.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active

    # Read header row (row 1)
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    headers_lc = [str(h).lower().strip() if h else "" for h in headers]

    def col(name: str) -> Optional[int]:
        for i, h in enumerate(headers_lc):
            if name in h:
                return i + 1
        return None

    c_cat    = col("catalog")
    c_lot    = col("lot")
    c_vendor = col("vendor") or col("supplier") or col("brand")
    c_marker = col("marker")
    c_fluor  = col("fluor")
    c_clone  = col("clone")
    c_iso    = col("isotype")
    c_conc   = col("concentration")
    c_exp    = col("expir") or col("expiry")
    c_dil    = col("dilution")

    # Build mapping: header_name → (col_idx, coa_field)
    fill_cols = {}
    for col_idx, hdr in enumerate(headers_lc, start=1):
        field = COL_FIELD_MAP.get(hdr)
        if field:
            fill_cols[col_idx] = field

    total_data_rows = ws.max_row - 1
    filled_count = 0
    skipped_count = 0
    error_count = 0

    progress_q.put({"type": "start", "total": total_data_rows})

    for row_idx in range(2, ws.max_row + 1):
        row_num = row_idx - 1  # 1-based data row number

        cat = str(ws.cell(row_idx, c_cat).value or "").strip() if c_cat else ""
        lot = str(ws.cell(row_idx, c_lot).value or "").strip() if c_lot else ""

        if not cat and not lot:
            progress_q.put({"type": "row", "row": row_num, "status": "skipped", "msg": "No catalog/lot"})
            skipped_count += 1
            continue

        # Check if all target fields are already filled
        all_filled = all(
            ws.cell(row_idx, ci).value not in (None, "", "NaN")
            for ci in fill_cols
            if ci is not None
        )
        if all_filled:
            progress_q.put({"type": "row", "row": row_num, "status": "skipped", "msg": f"{cat}/{lot} — already complete"})
            skipped_count += 1
            continue

        # Determine vendor for this row
        row_vendor = default_vendor
        if c_vendor:
            v_raw = str(ws.cell(row_idx, c_vendor).value or "").strip().lower()
            if v_raw:
                if "bd" in v_raw or "becton" in v_raw:
                    row_vendor = "bd"
                elif "milt" in v_raw:
                    row_vendor = "miltenyi"
                elif "bio" in v_raw:
                    row_vendor = "biolegend"

        progress_q.put({"type": "row", "row": row_num, "status": "fetching",
                        "msg": f"[{row_vendor}] Looking up {cat} / {lot}…"})

        coa = fetch_coa(row_vendor, cat, lot)

        if not coa["found"]:
            progress_q.put({"type": "row", "row": row_num, "status": "error",
                            "msg": f"{cat}/{lot} — CoA not found"})
            error_count += 1
        else:
            # Write each mapped field (only if cell is currently empty)
            written = []
            for ci, field in fill_cols.items():
                cell = ws.cell(row_idx, ci)
                if cell.value in (None, "", "NaN") and coa.get(field):
                    cell.value = coa[field]
                    written.append(field)

            # Highlight filled cells in light purple so the user can review
            fill_color = PatternFill("solid", fgColor="EDE1F7")
            for ci, _ in fill_cols.items():
                cell = ws.cell(row_idx, ci)
                if cell.value:
                    cell.fill = fill_color

            filled_count += 1
            progress_q.put({"type": "row", "row": row_num, "status": "ok",
                            "msg": f"{cat}/{lot} — {coa.get('product_name','') or 'filled'} ({', '.join(written) or 'no new data'})"})

        # Brief pause to be a polite HTTP client
        time.sleep(0.3)

    # Summary
    progress_q.put({
        "type": "done",
        "filled": filled_count,
        "skipped": skipped_count,
        "errors": error_count,
    })

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


# ════════════════════════════════════════════════════════════════════
# FLASK ROUTES
# ════════════════════════════════════════════════════════════════════

# In-memory store for batch jobs
_jobs: dict[str, dict] = {}   # job_id → {"status","result_bytes","q"}


@app.route("/")
def index():
    return HTML_PAGE, 200, {"Content-Type": "text/html; charset=utf-8"}


@app.errorhandler(Exception)
def _api_uncaught(e):
    """Last-resort handler so any backend exception comes back as JSON,
    not an HTML 500 page (which would break the frontend's resp.json())."""
    import traceback
    tb = traceback.format_exc()
    logger.error(f"Unhandled exception in {request.path}:\n{tb}")
    # Only override JSON for our API routes; let static/HTML routes 500 normally.
    if request.path.startswith("/api/"):
        return jsonify({
            "found": False,
            "error": f"Server error: {type(e).__name__}: {e}",
            "_traceback": tb.splitlines()[-5:],  # last 5 lines for diagnostics
        }), 500
    raise e


@app.route("/api/lookup", methods=["POST"])
def api_lookup():
    try:
        data = request.get_json(force=True, silent=True) or {}
        raw    = (data.get("barcode")    or "").strip()
        cat    = (data.get("catalog_no") or "").strip()
        lot    = (data.get("lot_no")     or "").strip()
        vendor = (data.get("vendor")     or "").strip().lower() or "biolegend"

        binfo = {}
        if raw:
            try:
                binfo = parse_barcode(raw, vendor) or {}
            except Exception as e:
                logger.warning(f"parse_barcode error: {e}")
                binfo = {}
            cat = cat or binfo.get("catalog_no", "")
            lot = lot or binfo.get("lot_no",     "")

        if not cat and not lot:
            return jsonify({"error": "No catalog or lot number could be extracted."}), 400

        coa = fetch_coa(vendor, cat, lot) or {}
        coa["barcode_info"] = binfo
        coa["vendor"]       = vendor
        return jsonify(coa)
    except Exception as e:
        import traceback
        logger.error(f"api_lookup error: {traceback.format_exc()}")
        return jsonify({
            "found": False,
            "error": f"Server error: {type(e).__name__}: {e}",
            "vendor": (data.get("vendor") if isinstance(data, dict) else "") or "biolegend",
            "fallback_url": "",
        }), 500


@app.route("/api/batch/start", methods=["POST"])
def api_batch_start():
    """Upload xlsx and start background processing. Returns job_id."""
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    f = request.files["file"]
    file_bytes = f.read()
    default_vendor = (request.form.get("vendor", "") or "").strip().lower() or "biolegend"

    import uuid
    job_id = str(uuid.uuid4())[:8]
    q = queue.Queue()
    _jobs[job_id] = {"status": "running", "result_bytes": None, "q": q}

    def worker():
        try:
            result = process_spreadsheet(file_bytes, q, default_vendor)
            _jobs[job_id]["result_bytes"] = result
            _jobs[job_id]["status"] = "done"
        except Exception as e:
            logger.exception("Batch worker error")
            _jobs[job_id]["status"] = "error"
            q.put({"type": "fatal", "msg": str(e)})

    t = threading.Thread(target=worker, daemon=True)
    t.start()
    return jsonify({"job_id": job_id})


@app.route("/api/batch/progress/<job_id>")
def api_batch_progress(job_id: str):
    """SSE stream of progress events for a batch job."""
    if job_id not in _jobs:
        return jsonify({"error": "Job not found"}), 404

    def event_stream() -> Generator[str, None, None]:
        job = _jobs[job_id]
        q = job["q"]
        while True:
            try:
                msg = q.get(timeout=30)
            except queue.Empty:
                yield "data: {\"type\":\"ping\"}\n\n"
                if job["status"] in ("done", "error"):
                    break
                continue

            yield f"data: {json.dumps(msg)}\n\n"
            if msg.get("type") in ("done", "fatal"):
                break

    return Response(event_stream(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.route("/api/batch/download/<job_id>")
def api_batch_download(job_id: str):
    job = _jobs.get(job_id)
    if not job or job["status"] != "done" or not job["result_bytes"]:
        return jsonify({"error": "Not ready or job not found"}), 404
    return send_file(
        io.BytesIO(job["result_bytes"]),
        as_attachment=True,
        download_name="CoA_Filled_Inventory.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/export_xlsx", methods=["POST"])
def api_export_xlsx():
    """Accept JSON array of CoA scan records, return a formatted .xlsx file."""
    records = request.get_json(force=True) or []
    if not records:
        return jsonify({"error": "No data provided"}), 400

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CoA Scan Results"

    col_keys = [
        "vendor",
        "catalog_no", "lot_no", "product_name", "marker", "fluor",
        "clone", "isotype", "concentration", "expiry_date",
        "applications", "storage", "usage", "scan_time",
    ]
    col_labels = [
        "Vendor",
        "Catalog #", "Lot #", "Product Name", "Marker", "Fluor",
        "Clone", "Isotype", "Concentration", "Expiry Date",
        "Applications", "Storage", "Usage", "Scanned",
    ]

    hdr_fill = PatternFill("solid", fgColor="6B4C9A")
    hdr_font = Font(bold=True, color="FFFFFF")
    row_fill = PatternFill("solid", fgColor="F4F1FA")

    # Header row
    for ci, label in enumerate(col_labels, 1):
        cell = ws.cell(1, ci, label)
        cell.fill = hdr_fill
        cell.font = hdr_font

    # Data rows
    for ri, rec in enumerate(records, 2):
        for ci, key in enumerate(col_keys, 1):
            ws.cell(ri, ci, rec.get(key, "") or "")
        if ri % 2 == 0:
            for ci in range(1, len(col_keys) + 1):
                ws.cell(ri, ci).fill = row_fill

    # Auto-fit column widths
    for col_cells in ws.columns:
        width = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(width + 3, 55)

    ws.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(
        out,
        as_attachment=True,
        download_name="CoA_Scan_Results.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ════════════════════════════════════════════════════════════════════
# EMBEDDED HTML
# ════════════════════════════════════════════════════════════════════

HTML_PAGE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1.0"/>
<title>CoA Lookup — BioLegend · BD · Miltenyi</title>
<style>
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
:root{
  --purple:#6B4C9A;--purple2:#8B6BBF;--blue:#4A90D9;
  --light:#F4F1FA;--border:#DDD3F0;--text:#2D2340;--muted:#6E6E8E;
  --success:#28A745;--danger:#DC3545;--warn:#E8A000;--white:#FFF;
}
body{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,Arial,sans-serif;background:var(--light);color:var(--text);min-height:100vh}

/* Header */
header{background:linear-gradient(135deg,var(--purple) 0%,var(--purple2) 60%,var(--blue) 100%);color:var(--white);padding:18px 32px;display:flex;align-items:center;gap:14px;box-shadow:0 2px 8px rgba(0,0,0,.22)}
header .logo{font-size:1.9rem}
header h1{font-size:1.35rem;font-weight:700}
header p{font-size:.82rem;opacity:.85;margin-top:2px}

/* Tabs */
.tabs{display:flex;gap:0;border-bottom:2px solid var(--border);margin:0 0 20px}
.tab{padding:11px 24px;font-size:.92rem;font-weight:600;cursor:pointer;border-bottom:3px solid transparent;color:var(--muted);transition:color .15s}
.tab.active{color:var(--purple);border-bottom-color:var(--purple)}
.tab:hover:not(.active){color:var(--text)}
.tab-panel{display:none}
.tab-panel.active{display:block}

main{max-width:960px;margin:0 auto;padding:24px 20px 60px}
.card{background:var(--white);border-radius:12px;box-shadow:0 2px 8px rgba(107,76,154,.10);padding:22px;border:1px solid var(--border);margin-bottom:18px}
.card-title{font-size:.73rem;font-weight:700;text-transform:uppercase;letter-spacing:.08em;color:var(--purple);margin-bottom:14px}

/* Scan input */
.scan-hint{text-align:center;color:var(--muted);font-size:.87rem;margin-bottom:11px}
.scan-hint strong{color:var(--text)}
#barcode-input{width:100%;font-size:1rem;padding:11px 15px;border:2px solid var(--border);border-radius:8px;background:var(--light);font-family:monospace;outline:none;transition:border-color .2s}
#barcode-input:focus{border-color:var(--purple);box-shadow:0 0 0 3px rgba(107,76,154,.14)}
.input-row{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-top:11px}
.input-group{display:flex;flex-direction:column;gap:4px}
.input-group label{font-size:.75rem;font-weight:600;color:var(--muted)}
.input-group input{padding:8px 11px;border:1.5px solid var(--border);border-radius:7px;font-size:.93rem;outline:none;background:#FAFAFA}
.input-group input:focus{border-color:var(--purple)}

.btn{display:inline-flex;align-items:center;gap:7px;padding:9px 20px;border:none;border-radius:8px;font-size:.9rem;font-weight:600;cursor:pointer;transition:background .18s,transform .1s}
.btn:active{transform:scale(.98)}
.btn-primary{background:var(--purple);color:var(--white)}
.btn-primary:hover{background:#5A3D85}
.btn-success{background:var(--success);color:var(--white)}
.btn-success:hover{background:#1E8E3E}
.btn-clear{background:#EEE;color:var(--text)}
.btn-clear:hover{background:#DDD}
.btn-row{display:flex;gap:9px;margin-top:13px;flex-wrap:wrap}

/* Chips */
#parsed-bar{display:none;align-items:center;flex-wrap:wrap;gap:7px;margin-top:12px;padding:9px 13px;background:var(--light);border-radius:8px;border:1px solid var(--border)}
.chip{display:inline-flex;align-items:center;gap:5px;padding:3px 11px;border-radius:99px;font-size:.79rem;font-weight:600}
.chip-cat{background:#E8F0FD;color:#2A5DBF;border:1px solid #C0D4F8}
.chip-lot{background:#EDF7EE;color:#1D7D35;border:1px solid #B3E5BC}
.chip-date{background:#FFF3E0;color:#A0520F;border:1px solid #FFD7A0}
.chip-fmt{background:#F3EDF9;color:var(--purple);border:1px solid var(--border)}

/* Status */
#status,#batch-status{display:none;align-items:center;gap:9px;padding:11px 15px;border-radius:9px;font-size:.88rem;font-weight:500;margin-bottom:14px}
.status-loading{background:#EEF4FF;color:var(--blue);border:1px solid #C6DDF8}
.status-error{background:#FFF0F0;color:var(--danger);border:1px solid #F8C6C6}
.status-success{background:#EDFBF0;color:var(--success);border:1px solid #B3E5BC}
.spinner{width:17px;height:17px;border:2.5px solid currentColor;border-top-color:transparent;border-radius:50%;animation:spin .7s linear infinite;flex-shrink:0}
@keyframes spin{to{transform:rotate(360deg)}}

/* CoA result */
#results{display:none}
.product-header{border-bottom:2px solid var(--light);padding-bottom:14px;margin-bottom:16px}
.product-header h2{font-size:1.18rem;font-weight:700;color:var(--purple);line-height:1.3}
.product-meta{display:flex;flex-wrap:wrap;gap:7px;margin-top:7px}
.info-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(190px,1fr));gap:11px;margin-bottom:16px}
.info-item .label{font-size:.70rem;font-weight:700;text-transform:uppercase;letter-spacing:.05em;color:var(--muted)}
.info-item .value{font-size:.93rem;color:var(--text);font-weight:500;margin-top:1px}
.test-table{width:100%;border-collapse:collapse;font-size:.86rem}
.test-table thead tr{background:var(--purple);color:var(--white)}
.test-table th{padding:8px 13px;text-align:left;font-weight:600}
.test-table td{padding:7px 13px;border-bottom:1px solid var(--border)}
.test-table tbody tr:nth-child(even){background:var(--light)}

/* ── Material checkbox ───────────────────────────────────────────────── */
.m-chk{
  appearance:none;-webkit-appearance:none;
  width:18px;height:18px;min-width:18px;
  border:2px solid #9E9E9E;border-radius:3px;
  background:#fff;cursor:pointer;
  display:inline-flex;align-items:center;justify-content:center;
  transition:background .15s,border-color .15s;
  vertical-align:middle;position:relative;
}
.m-chk:checked{background:var(--purple);border-color:var(--purple);}
.m-chk:checked::after{
  content:'';display:block;
  width:5px;height:9px;
  border:2px solid #fff;border-top:none;border-left:none;
  transform:rotate(45deg) translate(-1px,-1px);
}
.m-chk:indeterminate{background:var(--purple);border-color:var(--purple);}
.m-chk:indeterminate::after{
  content:'';display:block;
  width:9px;height:2px;background:#fff;border-radius:1px;
}
.m-chk:hover:not(:checked):not(:indeterminate){border-color:var(--purple);}

/* ── Material trash button ───────────────────────────────────────────── */
.btn-trash{
  background:none;border:none;cursor:pointer;
  padding:4px;border-radius:4px;
  color:#9E9E9E;display:inline-flex;align-items:center;justify-content:center;
  transition:color .15s,background .15s;
}
.btn-trash:hover{color:#E53935;background:#fce8e6;}
.external-link{display:inline-flex;align-items:center;gap:5px;color:var(--blue);text-decoration:none;font-size:.85rem;font-weight:500;padding:5px 11px;border-radius:6px;border:1px solid var(--border);background:#F0F7FF}
.external-link:hover{background:#D8EEFF}
.source-note{font-size:.72rem;color:var(--muted);margin-top:11px}

/* ─── Batch panel ─────────────────────────────────────────────── */
.drop-zone{border:2.5px dashed var(--bl-border,var(--border));border-radius:12px;padding:38px 20px;text-align:center;cursor:pointer;transition:background .2s,border-color .2s;background:var(--light)}
.drop-zone:hover,.drop-zone.drag-over{background:#EDE6F7;border-color:var(--purple)}
.drop-zone input{display:none}
.drop-zone .dz-icon{font-size:2.6rem;margin-bottom:8px}
.drop-zone .dz-text{font-size:.92rem;color:var(--muted)}
.drop-zone .dz-text strong{color:var(--text)}
.drop-zone .dz-filename{font-size:.85rem;font-weight:600;color:var(--purple);margin-top:6px}

/* Progress log */
#progress-log{display:none;max-height:340px;overflow-y:auto;border:1px solid var(--border);border-radius:9px;font-size:.80rem;font-family:monospace;background:#FAFAFA}
.log-row{padding:5px 12px;border-bottom:1px solid #F0EAF8;display:flex;align-items:center;gap:8px}
.log-row:last-child{border-bottom:none}
.log-row.ok    .icon::before{content:"✅"}
.log-row.error .icon::before{content:"❌"}
.log-row.skip  .icon::before{content:"⏭"}
.log-row.fetch .icon::before{content:"🔄";animation:spin .8s linear infinite;display:inline-block}
.icon{font-size:.95rem;width:1.4em;text-align:center}

/* Progress bar */
.prog-wrap{background:#E8E0F3;border-radius:99px;height:10px;margin:10px 0;overflow:hidden}
.prog-bar{height:100%;background:var(--purple);border-radius:99px;transition:width .4s}

/* Summary badges */
.badge{display:inline-flex;align-items:center;gap:5px;padding:4px 12px;border-radius:99px;font-size:.81rem;font-weight:700}
.badge-ok{background:#EDF7EE;color:var(--success);border:1px solid #B3E5BC}
.badge-err{background:#FFF0F0;color:var(--danger);border:1px solid #F8C6C6}
.badge-skip{background:#FFF8EE;color:var(--warn);border:1px solid #FFE0A0}

@media(max-width:600px){.input-row{grid-template-columns:1fr}.info-grid{grid-template-columns:1fr 1fr}}
</style>
</head>
<body>
<header>
  <div class="logo">🔬</div>
  <div>
    <h1>Multi-Vendor CoA Lookup</h1>
    <p>BioLegend · BD · Miltenyi — scan barcodes or batch-process inventory to auto-fill Certificate of Analysis data</p>
  </div>
</header>

<main>

<!-- TABS -->
<div class="tabs">
  <div class="tab active" onclick="switchTab('scan')">🔍 Single Scan</div>
  <div class="tab"        onclick="switchTab('batch')">📊 Batch Process Spreadsheet</div>
</div>

<!-- ══════════ SINGLE SCAN TAB ══════════ -->
<div id="tab-scan" class="tab-panel active">

  <div class="card">
    <div class="card-title">📷 Barcode Scanner / Manual Entry</div>
    <p class="scan-hint"><strong>Pick a vendor</strong>, then scan or type the catalog &amp; lot.</p>

    <!-- Vendor picker -->
    <div class="input-row" style="grid-template-columns:1fr">
      <div class="input-group">
        <label>Vendor</label>
        <select id="vendor-input"
          style="width:100%;padding:8px 10px;border:1.5px solid #D1C4E9;border-radius:8px;font-size:.95rem;background:#fff;color:#333;cursor:pointer"
          onchange="onVendorChanged()">
          <option value="biolegend">BioLegend</option>
          <option value="bd">BD (Becton Dickinson)</option>
          <option value="miltenyi">Miltenyi Biotec</option>
        </select>
      </div>
    </div>

    <input id="barcode-input" type="text" autocomplete="off" spellcheck="false"
           placeholder="Scan barcode here — e.g.  344742 B402098 2026/03/05"
           style="margin-top:11px"/>
    <div id="parsed-bar">
      <span style="font-size:.75rem;color:var(--muted);font-weight:600">DETECTED:</span>
      <span id="chip-cat"  class="chip chip-cat"  style="display:none"></span>
      <span id="chip-lot"  class="chip chip-lot"  style="display:none"></span>
      <span id="chip-date" class="chip chip-date" style="display:none"></span>
      <span id="chip-fmt"  class="chip chip-fmt"  style="display:none"></span>
    </div>
    <div class="input-row">
      <div class="input-group">
        <label>Catalog Number</label>
        <input id="cat-input" type="text" placeholder="e.g. 344742" autocomplete="off"/>
      </div>
      <div class="input-group">
        <label>Lot Number</label>
        <input id="lot-input" type="text" placeholder="e.g. B402098" autocomplete="off"/>
      </div>
    </div>
    <div class="input-row">
      <div class="input-group">
        <label>Usage</label>
        <select id="usage-input" style="width:100%;padding:8px 10px;border:1.5px solid #D1C4E9;border-radius:8px;font-size:.95rem;background:#fff;color:#333;cursor:pointer">
          <option value="">— select —</option>
          <option value="Clinical">Clinical</option>
          <option value="Translational Immunology">Translational Immunology</option>
          <option value="Non-clinical development">Non-clinical development</option>
        </select>
      </div>
    </div>
    <div class="btn-row">
      <button class="btn btn-primary" onclick="(()=>{const v=document.getElementById('vendor-input').value;const r=barcodeInput.value.trim();if(r){const p=parseBarcode(r,v);showChips(p);doLookup(r,p.catalog_no,p.lot_no);}else{doLookup('',catInput.value.trim(),lotInput.value.trim());}})()">🔍 Look Up CoA</button>
      <button class="btn btn-clear"   onclick="clearScan()">✕ Clear</button>
    </div>
  </div>

  <div id="status" style="display:none"></div>
  <div class="card" id="results" style="display:none"></div>

  <!-- ── Scan History Table ─────────────────────────────────────────── -->
  <div class="card" id="scan-history-card" style="display:none">
    <div style="display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:8px;margin-bottom:14px">
      <div class="card-title" style="margin:0">📋 Scan History &nbsp;<span id="scan-count"
        style="background:var(--purple);color:#fff;border-radius:99px;padding:1px 9px;font-size:.75rem">0</span>
        <span id="scan-selected-label" style="display:none;margin-left:6px;color:var(--muted);font-size:.75rem;font-weight:400"></span>
      </div>
      <div style="display:flex;gap:8px;flex-wrap:wrap">
        <button class="btn btn-clear"   onclick="clearHistory()"  style="font-size:.8rem;padding:5px 12px">✕ Clear all</button>
        <button class="btn btn-success" onclick="exportCSV()"     style="font-size:.8rem;padding:5px 12px">⬇ Export CSV</button>
        <button class="btn btn-primary" onclick="exportXLSX()"    style="font-size:.8rem;padding:5px 12px">⬇ Export Excel</button>
      </div>
    </div>
    <div style="overflow-x:auto">
      <table id="scan-table" class="test-table">
        <thead>
          <tr>
            <th style="width:32px;text-align:center">
              <input type="checkbox" id="chk-all" class="m-chk" title="Select all"
                onchange="toggleAll(this.checked)">
            </th>
            <th>Vendor</th>
            <th>Cat #</th><th>Lot #</th><th>Marker</th><th>Fluor</th>
            <th>Clone</th><th>Isotype</th><th>Conc.</th><th>Expiry</th>
            <th>Usage</th><th>Time</th>
            <th style="width:36px"></th>
          </tr>
        </thead>
        <tbody id="scan-tbody"></tbody>
      </table>
    </div>
  </div>

</div><!-- end tab-scan -->

<!-- ══════════ BATCH TAB ══════════ -->
<div id="tab-batch" class="tab-panel">

  <div class="card">
    <div class="card-title">📂 Upload Inventory Spreadsheet</div>
    <p style="font-size:.87rem;color:var(--muted);margin-bottom:14px">
      Upload your <strong>.xlsx</strong> antibody inventory. The tool will look up the CoA for every row that has
      a <strong>Catalog #</strong> and <strong>Lot #</strong>, then fill in
      Marker, Fluor, Clone, Isotype, Concentration, Expiration Date, and Optimal Dilution.
      If your spreadsheet has a <strong>Vendor</strong> column, each row uses its own vendor;
      otherwise the default below is used for every row.
    </p>

    <div class="input-row" style="grid-template-columns:1fr;margin-bottom:11px">
      <div class="input-group">
        <label>Default vendor (used when the spreadsheet has no Vendor column)</label>
        <select id="batch-vendor-input"
          style="width:100%;padding:8px 10px;border:1.5px solid #D1C4E9;border-radius:8px;font-size:.95rem;background:#fff;color:#333;cursor:pointer">
          <option value="biolegend">BioLegend</option>
          <option value="bd">BD (Becton Dickinson)</option>
          <option value="miltenyi">Miltenyi Biotec</option>
        </select>
      </div>
    </div>

    <div class="drop-zone" id="drop-zone" onclick="document.getElementById('file-input').click()">
      <input type="file" id="file-input" accept=".xlsx,.xls" onchange="onFileSelected(this.files[0])"/>
      <div class="dz-icon">📄</div>
      <div class="dz-text"><strong>Click to choose</strong> or drag &amp; drop your .xlsx file here</div>
      <div class="dz-filename" id="dz-filename"></div>
    </div>

    <div class="btn-row" style="margin-top:14px">
      <button class="btn btn-primary" id="btn-start" onclick="startBatch()" disabled>▶ Start Batch Lookup</button>
    </div>
  </div>

  <!-- Batch status + progress -->
  <div id="batch-status" style="display:none"></div>

  <div class="card" id="batch-progress-card" style="display:none">
    <div class="card-title">⚙️ Processing</div>
    <div class="prog-wrap"><div class="prog-bar" id="prog-bar" style="width:0%"></div></div>
    <div id="prog-label" style="font-size:.78rem;color:var(--muted);margin-bottom:10px"></div>
    <div id="progress-log"></div>
  </div>

  <!-- Download card -->
  <div class="card" id="batch-done-card" style="display:none">
    <div class="card-title">✅ Done</div>
    <div id="batch-summary" style="display:flex;gap:9px;flex-wrap:wrap;margin-bottom:14px"></div>
    <a id="download-link" class="btn btn-success" href="#" download="CoA_Filled_Inventory.xlsx">⬇ Download Filled Spreadsheet</a>
  </div>

</div><!-- end tab-batch -->

</main>

<script>
// ─── TAB SWITCHING ────────────────────────────────────────────────────────────
function switchTab(name){
  document.querySelectorAll('.tab').forEach((t,i)=>t.classList.toggle('active',['scan','batch'][i]===name));
  document.querySelectorAll('.tab-panel').forEach((p,i)=>p.classList.toggle('active',['tab-scan','tab-batch'][i]==='tab-'+name));
  if(name==='scan') setTimeout(()=>document.getElementById('barcode-input').focus(),50);
}

// ─── BARCODE PARSER (client-side mirror) ─────────────────────────────────────
function parseBarcode(raw, vendor){
  raw=raw.trim();
  vendor=(vendor||'').toLowerCase();
  const r={raw,catalog_no:'',lot_no:'',scan_date:'',format:'unknown',vendor:vendor};
  if(!raw) return r;

  // Miltenyi: AI 91 (any 3-digit subcode) + 9-digit catalog + 10-digit lot (+ trailers)
  //   positions 0-1   = "91"
  //   positions 2-4   = 3-digit sub-code (656, 906, 746, 956, ...)
  //   positions 5-13  = 9-digit catalog (e.g. 130123710 → "130-123-710")
  //   positions 14-23 = 10-digit lot
  const miltenyi=(s)=>{
    const d=s.replace(/\D/g,'');
    if(!d.startsWith('91')||d.length<24) return false;
    const cat=d.substr(5,9), lot=d.substr(14,10);
    if(!/^\d{9}$/.test(cat) || !/^\d{10}$/.test(lot)) return false;
    r.catalog_no=cat; r.lot_no=lot;
    r.format='miltenyi_gs1'; r.vendor='miltenyi';
    return true;
  };
  // BD: 01 + 14-digit GTIN + (17 YYMMDD optional) + 10 + 7-char lot
  const bd=(s)=>{
    const d=s.replace(/\D/g,'');
    if(!d.startsWith('01')||d.length<16+9) return false;
    const gtin=d.substr(2,14); let pos=16;
    if(d.substr(pos,2)==='17'&&d.length>=pos+8){
      const y=d.substr(pos+2,2),mo=d.substr(pos+4,2),da=d.substr(pos+6,2);
      r.scan_date=`20${y}/${mo}/${da}`;
      pos+=8;
    }
    if(d.substr(pos,2)!=='10') return false;
    const lot=d.substr(pos+2,7);
    const cat=String(parseInt(gtin.substr(7,6)));
    r.catalog_no=cat; r.lot_no=lot; r.format='bd_gs1'; r.vendor='bd';
    return true;
  };
  // BioLegend standard: "344742 B402098 2026/03/05 ..."
  const biolegend=(s)=>{
    const parts=s.split(/\s+/);
    if(parts.length>=2&&/^\d{4,8}$/.test(parts[0])&&/^[A-Za-z]\d{4,}$/i.test(parts[1])){
      r.catalog_no=parts[0]; r.lot_no=parts[1].toUpperCase();
      r.format='biolegend_standard'; r.vendor='biolegend';
      if(parts.length>2) r.scan_date=parts.slice(2).join(' ');
      return true;
    }
    return false;
  };
  const gs1Parens=(s)=>{
    if(!/\(\d{2}\)/.test(s)) return false;
    r.format='gs1_parenthesized';
    const lot=s.match(/\(10\)([^(]+)/); if(lot) r.lot_no=lot[1].trim().toUpperCase();
    const gt=s.match(/\(01\)(\d{13,14})/); if(gt) r.catalog_no=String(parseInt(gt[1].slice(7,13)));
    return true;
  };

  // Try vendor-biased first, then autodetect
  if(vendor==='miltenyi'&&miltenyi(raw)) return r;
  if(vendor==='bd'&&bd(raw))             return r;
  if(vendor==='biolegend'){
    if(biolegend(raw)) return r;
    if(gs1Parens(raw)) return r;
  }
  if(miltenyi(raw))  return r;
  if(biolegend(raw)) return r;
  if(gs1Parens(raw)) return r;
  if(bd(raw))        return r;

  if(raw.includes('\t')){const[c,l]=raw.split('\t');r.catalog_no=c.trim();r.lot_no=(l||'').trim().toUpperCase();r.format='tab_separated';return r;}
  if(/^[A-Za-z]\d{4,}$/.test(raw)){r.lot_no=raw.toUpperCase();r.format='lot_only';return r;}
  if(/^\d{4,8}$/.test(raw)){r.catalog_no=raw;r.format='catalog_only';return r;}
  r.lot_no=raw;r.format='unknown_as_lot';return r;
}

// Update placeholder / labels when vendor changes
function onVendorChanged(){
  const v=document.getElementById('vendor-input').value;
  const placeholders={
    biolegend:'Scan barcode — e.g.  344742 B402098 2026/03/05',
    bd:       'Scan barcode — e.g.  0100382905669699172707311050508742400026',
    miltenyi: 'Scan barcode — e.g.  916561301237105250101832...',
  };
  const hints={
    biolegend:['e.g. 344742','e.g. B402098'],
    bd:       ['e.g. 566969','e.g. 5050874'],
    miltenyi: ['e.g. 130-123-710','e.g. 5250101832'],
  };
  document.getElementById('barcode-input').placeholder=placeholders[v]||'';
  document.getElementById('cat-input').placeholder=hints[v][0];
  document.getElementById('lot-input').placeholder=hints[v][1];
}

// ─── SINGLE SCAN ─────────────────────────────────────────────────────────────
const barcodeInput=document.getElementById('barcode-input');
const catInput=document.getElementById('cat-input');
const lotInput=document.getElementById('lot-input');
onVendorChanged();
barcodeInput.focus();

document.addEventListener('click',e=>{
  // Never steal focus away from interactive controls (select / option / label /
  // textarea). Previously we only guarded <input>/<button>/.drop-zone, which
  // meant clicking the Vendor / Usage dropdowns immediately refocused the
  // barcode field and snapped the menu shut.
  if(e.target.closest('input,button,select,option,label,textarea,.drop-zone,.tabs,a'))
    return;
  barcodeInput.focus();
});

// When barcode is scanned, pass values DIRECTLY to doLookup as params
// so there is zero reliance on DOM field state between scans.
barcodeInput.addEventListener('keydown',e=>{
  if(e.key==='Enter'){
    e.preventDefault();
    const r=barcodeInput.value.trim();
    if(!r) return;
    const v=document.getElementById('vendor-input').value;
    const p=parseBarcode(r,v);
    showChips(p);
    // Pass raw barcode + parsed cat/lot directly — no catInput/lotInput read
    doLookup(r, p.catalog_no, p.lot_no);
  }
});
barcodeInput.addEventListener('input',()=>{
  const r=barcodeInput.value.trim();
  const v=document.getElementById('vendor-input').value;
  if(r.length>4) showChips(parseBarcode(r,v));
  else document.getElementById('parsed-bar').style.display='none';
});

function showChips(p){
  const bar=document.getElementById('parsed-bar');
  bar.style.display=p.catalog_no||p.lot_no?'flex':'none';
  const s=(id,icon,label,cls,show)=>{
    const el=document.getElementById(id);
    el.style.display=show?'':'none';
    el.textContent=show?`${icon} ${label}`:'';
    el.className=`chip ${cls}`;
  };
  s('chip-cat','📦',`Cat# ${p.catalog_no}`,'chip-cat',!!p.catalog_no);
  s('chip-lot','🏷️',`Lot ${p.lot_no}`,'chip-lot',!!p.lot_no);
  s('chip-date','📅',p.scan_date,'chip-date',!!p.scan_date);
  s('chip-fmt','📊',p.format.replace(/_/g,' '),'chip-fmt',true);
}

// doLookup accepts explicit params; falls back to reading fields when called
// from the manual-entry button or cat/lot Enter key.
async function doLookup(rawBarcode, catalogNo, lotNo){
  const raw = rawBarcode !== undefined ? rawBarcode : barcodeInput.value.trim();
  const cat = catalogNo  !== undefined ? catalogNo  : catInput.value.trim();
  const lot = lotNo      !== undefined ? lotNo       : lotInput.value.trim();
  const vendor = document.getElementById('vendor-input').value || 'biolegend';
  if(!raw&&!cat&&!lot){setStatus('error','⚠️ Please scan a barcode or enter catalog / lot number.');return;}
  const vendorName = {biolegend:'BioLegend', bd:'BD', miltenyi:'Miltenyi'}[vendor] || vendor;
  setStatus('loading', `Fetching Certificate of Analysis from ${vendorName}…`);
  document.getElementById('results').style.display='none';
  try{
    const resp=await fetch('/api/lookup',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({barcode:raw, catalog_no:cat, lot_no:lot, vendor:vendor})});
    // Read response as text first, then try to JSON-parse so we can surface
    // useful info even when the server hits an unhandled exception page.
    const txt = await resp.text();
    let d;
    try { d = JSON.parse(txt); }
    catch(_){
      const snippet = (txt||'').replace(/\\s+/g,' ').slice(0,300);
      setStatus('error',`⚠️ Server returned non-JSON (HTTP ${resp.status}). First 300 chars: ${snippet}`);
      return;
    }
    if(d.error&&!d.found){setStatus('error',`⚠️ ${d.error}`);renderFallback(d);return;}
    setStatus('success','✅ Certificate of Analysis retrieved.');
    renderCoA(d);
    addToHistory(d);
    // Update display fields from the confirmed response data
    catInput.value = d.catalog_no || cat || '';
    lotInput.value = d.lot_no     || lot || '';
    // Clear barcode field and refocus for next scan
    barcodeInput.value='';
    document.getElementById('parsed-bar').style.display='none';
    barcodeInput.focus();
  }catch(err){setStatus('error',`⚠️ Network error: ${err.message}`);}
}

function setStatus(type,msg){
  const el=document.getElementById('status');
  el.style.display='flex';
  el.className=`status-${type}`;
  el.innerHTML=type==='loading'?`<div class="spinner"></div><span>${msg}</span>`:`<span>${msg}</span>`;
}

function renderCoA(d){
  const el=document.getElementById('results');
  el.style.display='block';
  const fi=(lbl,val)=>val?`<div class="info-item"><div class="label">${lbl}</div><div class="value">${h(val)}</div></div>`:'';
  const info=[
    fi('Catalog #',d.catalog_no), fi('Lot #',d.lot_no),
    fi('Marker / Target',d.marker), fi('Fluorochrome',d.fluor),
    fi('Clone',d.clone), fi('Isotype',d.isotype),
    fi('Host Species',d.host_species), fi('Reactivity',d.reactivity),
    fi('Concentration',d.concentration), fi('Optimal Dilution',d.optimal_dilution),
    fi('Expiry Date',d.expiry_date), fi('Storage',d.storage),
    fi('Formulation',d.formulation),
  ].filter(Boolean).join('');
  let testHtml='';
  if(d.tests&&d.tests.length>0){
    const hs=d.tests.some(t=>t.specification),hm=d.tests.some(t=>t.method);
    testHtml=`<div class="card-title" style="margin-top:18px">🧪 Test Results</div>
    <div style="overflow-x:auto"><table class="test-table">
    <thead><tr><th>Test / Parameter</th><th>Result</th>${hs?'<th>Specification</th>':''}${hm?'<th>Method</th>':''}</tr></thead>
    <tbody>${d.tests.map(t=>`<tr><td>${h(t.test)}</td><td>${h(t.result)}</td>${hs?`<td>${h(t.specification||'')}</td>`:''}${hm?`<td>${h(t.method||'')}</td>`:''}</tr>`).join('')}</tbody>
    </table></div>`;
  }
  const vendorLabel={biolegend:'BioLegend',bd:'BD',miltenyi:'Miltenyi'}[d.vendor]||'Vendor';
  const link=d.fallback_url?`<a class="external-link" href="${h(d.fallback_url)}" target="_blank">🌐 View on ${h(vendorLabel)}</a>`:'';
  el.innerHTML=`
  <div class="product-header">
    <h2>${h(d.product_name||(vendorLabel+' Product'))}</h2>
    <div class="product-meta">
      ${d.catalog_no?`<span class="chip chip-cat">📦 Cat# ${h(d.catalog_no)}</span>`:''}
      ${d.lot_no?`<span class="chip chip-lot">🏷️ Lot ${h(d.lot_no)}</span>`:''}
      ${link}
    </div>
  </div>
  ${info?`<div class="card-title">📋 Product Information</div><div class="info-grid">${info}</div>`:''}
  ${testHtml}
  ${d.source_url?`<p class="source-note">Source: <a href="${h(d.source_url)}" target="_blank">${h(d.source_url)}</a></p>`:''}
  ${d.error?`<p style="color:var(--muted);font-size:.8rem;margin-top:12px">ℹ️ ${h(d.error)}</p>`:''}`;
}

function renderFallback(d){
  const el=document.getElementById('results');el.style.display='block';
  const vendorLabel={biolegend:'BioLegend',bd:'BD',miltenyi:'Miltenyi'}[d.vendor]||'vendor';
  el.innerHTML=`<div class="card-title">⚠️ Manual Lookup Required</div>
  <p style="color:var(--muted);font-size:.88rem;margin-bottom:12px">Could not retrieve CoA automatically.</p>
  ${d.fallback_url?`<a class="external-link" href="${h(d.fallback_url)}" target="_blank">🌐 Open CoA on ${h(vendorLabel)} website</a>`:''}`;
}

function clearScan(){
  barcodeInput.value='';catInput.value='';lotInput.value='';
  document.getElementById('parsed-bar').style.display='none';
  document.getElementById('status').style.display='none';
  document.getElementById('results').style.display='none';
  barcodeInput.focus();
}
[catInput,lotInput].forEach(inp=>inp.addEventListener('keydown',e=>{
  if(e.key==='Enter') doLookup('', catInput.value.trim(), lotInput.value.trim());
}));

// ─── BATCH PROCESSING ────────────────────────────────────────────────────────
let batchFile=null;

const dropZone=document.getElementById('drop-zone');
dropZone.addEventListener('dragover',e=>{e.preventDefault();dropZone.classList.add('drag-over')});
dropZone.addEventListener('dragleave',()=>dropZone.classList.remove('drag-over'));
dropZone.addEventListener('drop',e=>{
  e.preventDefault();dropZone.classList.remove('drag-over');
  const f=e.dataTransfer.files[0];
  if(f) onFileSelected(f);
});

function onFileSelected(file){
  if(!file) return;
  batchFile=file;
  document.getElementById('dz-filename').textContent=`📄 ${file.name}  (${(file.size/1024).toFixed(1)} KB)`;
  document.getElementById('btn-start').disabled=false;
}

async function startBatch(){
  if(!batchFile){alert('Please choose a file first.');return;}
  document.getElementById('btn-start').disabled=true;
  document.getElementById('batch-done-card').style.display='none';
  document.getElementById('progress-log').innerHTML='';

  const fd=new FormData();
  fd.append('file',batchFile);
  fd.append('vendor', document.getElementById('batch-vendor-input').value || 'biolegend');

  setBatchStatus('loading','Uploading file and starting batch lookup…');
  document.getElementById('batch-progress-card').style.display='block';
  document.getElementById('progress-log').style.display='block';

  let resp;
  try{resp=await fetch('/api/batch/start',{method:'POST',body:fd});}
  catch(e){setBatchStatus('error',`Upload failed: ${e.message}`);return;}

  const {job_id,error}=await resp.json();
  if(error){setBatchStatus('error',`Error: ${error}`);return;}

  // Open SSE stream
  const es=new EventSource(`/api/batch/progress/${job_id}`);
  let total=0,done=0;

  es.onmessage=e=>{
    const msg=JSON.parse(e.data);
    if(msg.type==='ping') return;

    if(msg.type==='start'){
      total=msg.total;
      setBatchStatus('loading',`Processing ${total} rows…`);
    } else if(msg.type==='row'){
      done++;
      const pct=total>0?Math.round(done/total*100):0;
      document.getElementById('prog-bar').style.width=pct+'%';
      document.getElementById('prog-label').textContent=`Row ${done} / ${total}  (${pct}%)`;
      addLogRow(msg.status,msg.msg||'');
    } else if(msg.type==='done'){
      es.close();
      setBatchStatus('success','✅ Batch complete!');
      showBatchSummary(msg,job_id);
    } else if(msg.type==='fatal'){
      es.close();
      setBatchStatus('error',`Fatal error: ${msg.msg}`);
    }
  };
  es.onerror=()=>{es.close();setBatchStatus('error','Connection lost. Check server logs.');};
}

function setBatchStatus(type,msg){
  const el=document.getElementById('batch-status');
  el.style.display='flex';
  el.className=`status-${type}`;
  el.innerHTML=type==='loading'?`<div class="spinner"></div><span>${msg}</span>`:`<span>${msg}</span>`;
}

function addLogRow(status,msg){
  const log=document.getElementById('progress-log');
  const cls=status==='ok'?'ok':status==='error'?'error':status==='fetching'?'fetch':'skip';
  const row=document.createElement('div');
  row.className=`log-row ${cls}`;
  row.innerHTML=`<span class="icon"></span><span>${h(msg)}</span>`;
  log.appendChild(row);
  log.scrollTop=log.scrollHeight;
}

function showBatchSummary(msg,job_id){
  document.getElementById('batch-done-card').style.display='block';
  document.getElementById('batch-summary').innerHTML=`
    <span class="badge badge-ok">✅ Filled: ${msg.filled}</span>
    <span class="badge badge-skip">⏭ Skipped: ${msg.skipped}</span>
    <span class="badge badge-err">❌ Errors: ${msg.errors}</span>`;
  document.getElementById('download-link').href=`/api/batch/download/${job_id}`;
  document.getElementById('btn-start').disabled=false;
}

function h(s){
  return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');
}

// ─── SCAN HISTORY ─────────────────────────────────────────────────────────────
let scanHistory=[];   // array of {entry, rowEl} objects
const HISTORY_COLS=['vendor','catalog_no','lot_no','product_name','marker','fluor','clone','isotype','concentration','expiry_date','applications','storage','usage','scan_time'];
const HISTORY_HDRS=['Vendor','Catalog #','Lot #','Product Name','Marker','Fluor','Clone','Isotype','Concentration','Expiry Date','Applications','Storage','Usage','Scanned'];
const VENDOR_LABEL={biolegend:'BioLegend',bd:'BD',miltenyi:'Miltenyi'};

function _checkedRows(){
  return scanHistory.filter(item=>item.rowEl.querySelector('.row-chk').checked);
}

function _updateSelectionLabel(){
  const total=scanHistory.length;
  const sel=_checkedRows().length;
  const lbl=document.getElementById('scan-selected-label');
  if(sel>0&&sel<total){
    lbl.style.display='';
    lbl.textContent=`(${sel} of ${total} selected for export)`;
  } else {
    lbl.style.display='none';
  }
  // Sync select-all checkbox state
  const chkAll=document.getElementById('chk-all');
  if(chkAll){
    chkAll.indeterminate = sel>0 && sel<total;
    chkAll.checked = !chkAll.indeterminate && sel===total && total>0;
    // Drive the CSS ::after via a data attr since indeterminate is JS-only
    chkAll.dataset.indeterminate = chkAll.indeterminate ? '1' : '';
  }
}

function toggleAll(checked){
  scanHistory.forEach(item=>{
    item.rowEl.querySelector('.row-chk').checked=checked;
  });
  _updateSelectionLabel();
}

function addToHistory(d){
  const vendorKey = (d.vendor || document.getElementById('vendor-input').value || 'biolegend').toLowerCase();
  const entry={
    vendor:       VENDOR_LABEL[vendorKey] || vendorKey,
    catalog_no:   d.catalog_no||'',
    lot_no:       d.lot_no||'',
    product_name: d.product_name||'',
    marker:       d.marker||'',
    fluor:        d.fluor||'',
    clone:        d.clone||'',
    isotype:      d.isotype||'',
    concentration:d.concentration||'',
    expiry_date:  d.expiry_date||'',
    applications: d.applications||'',
    storage:      d.storage||'',
    usage:        document.getElementById('usage-input').value,
    scan_time:    new Date().toLocaleTimeString(),
  };
  const idx=scanHistory.length;
  const tr=document.createElement('tr');
  tr.innerHTML=
    `<td style="text-align:center"><input type="checkbox" class="row-chk m-chk" checked
        onchange="_updateSelectionLabel()"></td>`+
    `<td><span class="chip chip-fmt">${h(entry.vendor)}</span></td>`+
    `<td>${h(entry.catalog_no)}</td><td>${h(entry.lot_no)}</td>`+
    `<td>${h(entry.marker)}</td><td>${h(entry.fluor)}</td>`+
    `<td>${h(entry.clone)}</td><td>${h(entry.isotype)}</td>`+
    `<td>${h(entry.concentration)}</td><td>${h(entry.expiry_date)}</td>`+
    `<td>${h(entry.usage)}</td>`+
    `<td style="color:var(--muted);font-size:.8rem">${h(entry.scan_time)}</td>`+
    `<td style="text-align:center">
       <button class="btn-trash" title="Delete this row" onclick="deleteRow(${idx})">
         <svg xmlns="http://www.w3.org/2000/svg" width="16" height="16" viewBox="0 0 24 24" fill="currentColor">
           <path d="M6 19c0 1.1.9 2 2 2h8c1.1 0 2-.9 2-2V7H6v12zM19 4h-3.5l-1-1h-5l-1 1H5v2h14V4z"/>
         </svg>
       </button>
     </td>`;
  scanHistory.push({entry, rowEl:tr});
  document.getElementById('scan-tbody').appendChild(tr);
  document.getElementById('scan-count').textContent=scanHistory.length;
  document.getElementById('scan-history-card').style.display='block';
  _updateSelectionLabel();
  if(scanHistory.length>1)
    document.getElementById('scan-history-card').scrollIntoView({behavior:'smooth',block:'nearest'});
}

function deleteRow(idx){
  const item=scanHistory[idx];
  if(!item||!item.rowEl) return;
  const label=item.entry.catalog_no
    ? `Cat# ${item.entry.catalog_no} / Lot ${item.entry.lot_no}`
    : 'this entry';
  if(!confirm(`Delete ${label} from scan history?`)) return;
  item.rowEl.remove();
  // Mark as deleted (keep index stable; filter on export)
  item._deleted=true;
  scanHistory=scanHistory.filter(i=>!i._deleted);
  // Rebind delete button indices after removal
  document.querySelectorAll('#scan-tbody tr').forEach((tr,i)=>{
    const btn=tr.querySelector('button[title="Delete this row"]');
    if(btn) btn.setAttribute('onclick',`deleteRow(${i})`);
  });
  document.getElementById('scan-count').textContent=scanHistory.length;
  if(!scanHistory.length) document.getElementById('scan-history-card').style.display='none';
  _updateSelectionLabel();
}

function clearHistory(){
  if(!scanHistory.length) return;
  if(!confirm(`Clear all ${scanHistory.length} entr${scanHistory.length===1?'y':'ies'} from scan history?`)) return;
  scanHistory=[];
  document.getElementById('scan-tbody').innerHTML='';
  document.getElementById('scan-count').textContent=0;
  document.getElementById('scan-history-card').style.display='none';
}

function _rowsToExport(){
  const checked=_checkedRows();
  // If nothing is explicitly unchecked (all checked or none exist) export all
  const rows = checked.length>0 ? checked : scanHistory;
  if(!rows.length){alert('No scans to export.');return null;}
  return rows.map(i=>i.entry);
}

function exportCSV(){
  const rows=_rowsToExport(); if(!rows) return;
  const data=[HISTORY_HDRS,...rows.map(r=>HISTORY_COLS.map(c=>JSON.stringify(r[c]||'')))];
  const csv=data.map(r=>r.join(',')).join('\r\n');
  const a=document.createElement('a');
  a.href=URL.createObjectURL(new Blob([csv],{type:'text/csv;charset=utf-8;'}));
  a.download='CoA_Scan_Results.csv';
  a.click();
}

async function exportXLSX(){
  const rows=_rowsToExport(); if(!rows) return;
  try{
    const resp=await fetch('/api/export_xlsx',{
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body:JSON.stringify(rows)
    });
    if(!resp.ok){alert('Export failed: '+resp.statusText);return;}
    const blob=await resp.blob();
    const a=document.createElement('a');
    a.href=URL.createObjectURL(blob);
    a.download='CoA_Scan_Results.xlsx';
    a.click();
  }catch(e){alert('Export error: '+e.message);}
}
</script>
</body>
</html>
"""

if __name__ == "__main__":
    import sys
    port = int(sys.argv[1]) if len(sys.argv) > 1 else 5050
    print(f"\n🧬  BioLegend CoA Lookup")
    print(f"    Open  http://localhost:{port}  in your browser\n")
    app.run(debug=False, port=port, host="0.0.0.0")
